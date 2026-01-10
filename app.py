# Bibliotecas padrão
import io
import os
import sys
import time
from datetime import date, datetime, timedelta
from decimal import Decimal, getcontext
from typing import Dict, List, Optional, Tuple, Union

# Bibliotecas de terceiros
import streamlit as st
import pandas as pd
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
import psycopg2
import psycopg2.extras

# Adiciona o diretório pai ao path para importar módulos locais
sys.path.append('..')

# Módulos locais
from pdfgenerator import gerar_contrato_pdf, gerar_recibo_pdf
from pdfgenerator import STATUS_CARRO, STATUS_CLIENTE, STATUS_RESERVA
from init_db import init_db_production, check_db_health

from auth_utils import (
    verify_credentials,
    get_current_user,
    require_role,
    logout as auth_logout,
)
from auth_manager import auth_manager as supabase_auth_manager, USER_ROLES
from db_utils import run_query, run_query_dataframe, get_db_connection


def get_reservas_entrega():
    """Busca reservas aguardando entrega com dados completos em uma única query"""
    query = """
        SELECT 
            r.id, r.carro_id, r.cliente_id, r.data_inicio, r.data_fim, r.total_diarias, r.adiantamento,
            c.nome AS cliente_nome, c.cpf, c.telefone,
            carros.modelo, carros.placa, carros.km_atual, carros.diaria
        FROM reservas r
        JOIN clientes c ON r.cliente_id = c.id
        JOIN carros ON r.carro_id = carros.id
        WHERE r.status = 'Ativa' 
        AND r.reserva_status = 'Reservada'
        AND carros.status != 'Locado'
        ORDER BY r.data_inicio ASC
    """
    return run_query_dataframe(query)


def get_relatorio_ocupacao_mensal(ano_selecionado, mes_selecionado):
    """Busca dados do relatório de ocupação mensal em uma única query"""
    primeiro_dia_mes = date(ano_selecionado, mes_selecionado, 1)
    if mes_selecionado == 12:
        ultimo_dia_mes = date(ano_selecionado, mes_selecionado, 31)
    else:
        ultimo_dia_mes = date(ano_selecionado, mes_selecionado + 1, 1) - timedelta(days=1)
    
    query = """
        WITH carros_ativos AS (
            SELECT id, modelo, placa
            FROM carros 
            WHERE status != %s
        ),
        reservas_periodo AS (
            SELECT carro_id, data_inicio, data_fim, reserva_status
            FROM reservas
            WHERE (reserva_status IN ('Reservada', 'Locada', 'Finalizada'))
            AND (data_inicio <= %s AND data_fim >= %s)
        )
        SELECT 
            ca.id, ca.modelo, ca.placa,
            rp.carro_id, rp.data_inicio, rp.data_fim, rp.reserva_status
        FROM carros_ativos ca
        LEFT JOIN reservas_periodo rp ON ca.id = rp.carro_id
        ORDER BY ca.modelo, ca.placa
    """
    
    return run_query_dataframe(query, (STATUS_CARRO['EXCLUIDO'], ultimo_dia_mes, primeiro_dia_mes))


def get_dashboard_data():
    """Busca todos os dados do dashboard em uma única query otimizada"""
    query = """
        WITH locados AS (
            SELECT COUNT(*) as total
            FROM reservas r
            WHERE r.status = 'Ativa' AND r.reserva_status = 'Locada'
        ),
        reservados AS (
            SELECT COUNT(*) as total
            FROM reservas r
            WHERE r.status = 'Ativa' AND r.reserva_status = 'Reservada'
        ),
        faturamento AS (
            SELECT COALESCE(SUM(valor_total), 0) as total
            FROM reservas
            WHERE reserva_status = 'Finalizada'
            AND data_fim BETWEEN DATE_TRUNC('month', CURRENT_DATE) 
                             AND DATE_TRUNC('month', CURRENT_DATE) + INTERVAL '1 month'
        ),
        devolucoes_hoje AS (
            SELECT COUNT(*) as total
            FROM reservas
            WHERE reserva_status = 'Locada'
            AND data_fim::date = CURRENT_DATE
        )
        SELECT 
            (SELECT COUNT(*) FROM carros WHERE status != %s) as total_carros,
            (SELECT total FROM locados) as carros_locados,
            (SELECT total FROM reservados) as carros_reservados,
            (SELECT total FROM faturamento) as faturamento_mensal,
            (SELECT total FROM devolucoes_hoje) as devolucoes_hoje
    """
    return run_query_dataframe(query, (STATUS_CARRO['EXCLUIDO'],)).iloc[0].to_dict()


# --- CONFIGURAÇÃO INICIAL E DESIGN SYSTEM ---
st.set_page_config(page_title="Locadora Iguacu Veiculos", layout="wide", page_icon="🚗")

THEME_COLORS = {
    "primary": "#1D4ED8",
    "primary_soft": "#E0E7FF",
    "accent": "#F97316",
    "gray_900": "#0F172A",
    "gray_600": "#475569",
    "gray_200": "#E2E8F0",
    "background": "#F6F7FB",
}


def inject_global_styles():
    st.markdown(
        f"""
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600&display=swap');

            :root {{
                --brand-primary: {THEME_COLORS["primary"]};
                --brand-primary-soft: {THEME_COLORS["primary_soft"]};
                --brand-accent: {THEME_COLORS["accent"]};
                --brand-gray-900: {THEME_COLORS["gray_900"]};
                --brand-gray-600: {THEME_COLORS["gray_600"]};
                --brand-gray-200: {THEME_COLORS["gray_200"]};
                --brand-bg: {THEME_COLORS["background"]};
            }}

            html, body, [class*="stApp"] {{
                font-family: 'Space Grotesk', 'Segoe UI', sans-serif;
                background: var(--brand-bg);
                color: var(--brand-gray-900);
            }}

            .main .block-container {{
                padding-top: 1.5rem;
                padding-bottom: 4rem;
                padding-left: 2.5rem;
                padding-right: 2.5rem;
            }}

            section[data-testid="stSidebar"] {{
                background: white;
                border-right: 1px solid var(--brand-gray-200);
            }}

            section[data-testid="stSidebar"] .block-container {{
                padding: 1.5rem 1.25rem 2.5rem;
            }}

            .sidebar-logo {{
                display: flex;
                flex-direction: column;
                gap: 0.25rem;
                margin-bottom: 1.25rem;
            }}

            .sidebar-logo__title {{
                font-size: 1.2rem;
                font-weight: 600;
                color: var(--brand-gray-900);
            }}

            .sidebar-logo__subtitle {{
                color: var(--brand-gray-600);
                font-size: 0.9rem;
            }}

            .stRadio > label {{
                font-weight: 500 !important;
            }}

            .stRadio [role="radiogroup"] {{
                display: flex;
                flex-direction: column;
                gap: 0.35rem;
            }}

            .stRadio [role="radio"] span {{
                padding: 0.55rem 0.75rem;
                border-radius: 0.6rem;
                border: 1px solid transparent;
            }}

            .stRadio [aria-checked="true"] span {{
                background: var(--brand-primary-soft);
                border-color: var(--brand-primary);
                color: var(--brand-primary);
            }}

            .section-header {{
                display: flex;
                justify-content: space-between;
                align-items: flex-start;
                gap: 0.5rem;
                margin-bottom: 1rem;
                padding-bottom: 0.5rem;
                border-bottom: 1px solid var(--brand-gray-200);
            }}

            .section-header__title {{
                font-size: 1.35rem;
                font-weight: 600;
                color: var(--brand-gray-900);
            }}

            .section-header__subtitle {{
                font-size: 0.95rem;
                color: var(--brand-gray-600);
                margin-top: 0.25rem;
            }}

            .breadcrumb {{
                font-size: 0.85rem;
                color: var(--brand-gray-600);
                letter-spacing: 0.02em;
                text-transform: uppercase;
            }}

            .stButton>button, .stDownloadButton>button {{
                border-radius: 0.75rem;
                padding: 0.6rem 1.4rem;
                border: 1px solid transparent;
                background: var(--brand-primary);
                color: white;
                font-weight: 600;
                transition: all 0.2s ease;
            }}

            .stButton>button:hover, .stDownloadButton>button:hover {{
                box-shadow: 0 10px 20px rgba(29,78,216,0.25);
                transform: translateY(-1px);
            }}

            .metric-card {{
                padding: 1rem 1.2rem;
                background: white;
                border-radius: 1rem;
                border: 1px solid rgba(15, 23, 42, 0.06);
                box-shadow: 0 6px 16px rgba(15, 23, 42, 0.05);
            }}

            .metric-card__label {{
                font-size: 0.85rem;
                text-transform: uppercase;
                color: var(--brand-gray-600);
                margin-bottom: 0.2rem;
            }}

            .metric-card__value {{
                font-size: 1.65rem;
                font-weight: 600;
                color: var(--brand-gray-900);
            }}

            .metric-card__helper {{
                font-size: 0.85rem;
                color: var(--brand-gray-600);
                margin-top: 0.15rem;
            }}

            .section-divider {{
                margin: 2.5rem 0 1.5rem;
                border: none;
                border-top: 1px dashed var(--brand-gray-200);
            }}

            .stTextInput>div>div>input,
            .stSelectbox>div>div>select,
            .stTextArea>div>textarea,
            .stNumberInput>div>div>input {{
                border-radius: 0.65rem;
                border: 1px solid var(--brand-gray-200);
                padding: 0.55rem 0.75rem;
            }}

            .stTabs [data-baseweb="tab-list"] {{
                gap: 0.25rem;
            }}

            .stTabs [role="tab"] {{
                padding: 0.45rem 0.9rem;
                border-radius: 999px;
            }}

            .stTable, .stDataFrame {{
                border-radius: 1rem;
                overflow: hidden;
                box-shadow: 0 12px 24px rgba(15, 23, 42, 0.05);
            }}
        </style>
        """,
        unsafe_allow_html=True,
    )
    
   
def render_section_header(title: str, subtitle: Optional[str] = None, icon: str = "", trail: Optional[List[str]] = None):
    breadcrumb_items = ["Início"]
    if trail:
        breadcrumb_items.extend(trail)
    breadcrumb_text = " / ".join(breadcrumb_items)
    st.markdown(f"<div class='breadcrumb'>{breadcrumb_text}</div>", unsafe_allow_html=True)
    icon_markup = f"<span style='margin-right:0.5rem'>{icon}</span>" if icon else ""
    subtitle_markup = f"<div class='section-header__subtitle'>{subtitle}</div>" if subtitle else ""
    st.markdown(
        f"""
        <div class='section-header'>
            <div>
                <div class='section-header__title'>{icon_markup}{title}</div>
                {subtitle_markup}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

# --- AUTENTICAÇÃO SIMPLIFICADA ---
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False


def check_password():
    """Exibe o formulário de login e valida as credenciais via Supabase."""

    def password_entered():
        user = verify_credentials(
            st.session_state.get("username", "").strip(),
            st.session_state.get("password", ""),
        )
        if user and user.get("is_active", False):
            st.session_state["user"] = user
            st.session_state["authenticated"] = True
            st.session_state["password_correct"] = True
            st.session_state.pop("password", None)
        else:
            st.session_state["password_correct"] = False

    # Estilo para centralização
    st.markdown("""
        <style>
            .login-container {
                max-width: 400px;
                margin: 0 auto;
                padding: 2rem;
            }
            .stTextInput>div>div>input {
                padding: 0.5rem;
            }
        </style>
    """, unsafe_allow_html=True)

    # Container centralizado
    with st.container():
        col1, col2, col3 = st.columns([1, 2, 1])
        
        with col2:
            st.title("🔐 Locadora Iguacu Veículos")
            st.caption("Informe suas credenciais para acessar o sistema.")

            with st.form("login_form"):
                st.text_input("Email", key="username", placeholder="seu@email.com")
                st.text_input("Senha", type="password", key="password", placeholder="••••••••")
                
                submitted = st.form_submit_button("Entrar", type="primary", width='stretch')

                if submitted:
                    if not st.session_state.get("username") or not st.session_state.get("password"):
                        st.error("Por favor, preencha usuário e senha.")
                    else:
                        password_entered()

    if st.session_state.get("password_correct") is False:
        st.error("😕 Usuário ou senha incorretos. Tente novamente.")

    return st.session_state.get("authenticated", False)


if not st.session_state.get("authenticated", False):
    if not check_password():
        st.stop()
    else:
        st.rerun()

current_user = get_current_user()
if not current_user:
    st.error("Erro ao carregar informações do usuário. Faça login novamente.")
    st.session_state.authenticated = False
    st.stop()


def logout_user():
    """Finaliza a sessão atual e retorna para o formulário de login."""
    auth_logout()
    st.session_state.pop("menu_initialized", None)
    st.rerun()


# --- FUNÇÃO PRINCIPAL ---
def main():
    """
    Função principal da aplicação com controle de acesso baseado em função.
    """
    require_role(['admin', 'manager', 'viewer'])
    st.header("Bem-vindo ao Sistema de Locação")
    st.write(f"Olá, {current_user.get('full_name', 'Usuário')}!")
    
    # Conteúdo protegido aqui
    if st.button("Sair"):
        logout_user()


# --- INICIALIZAÇÃO DO BANCO DE DADOS ---
# Verificar e inicializar banco para produção
db_health = check_db_health()
if not db_health['healthy']:
    init_db_production()

# --- FUNÇÕES DE FORMATAÇÃO E UTILIDADE ---

def formatar_moeda(valor):
    """Formata um valor float para a moeda brasileira (R$ 0.000,00)."""
    if valor is None:
        valor = 0.0
    # Garante que o separador decimal seja vírgula e o milhar seja ponto
    return f"R$ {float(valor):,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


# --- FUNÇÕES DE DISPONIBILIDADE DE VEÍCULOS ---

def get_available_vehicles(data_inicio, data_fim, permitir_dia_devolucao=False):
    """
    Retorna veículos disponíveis para o período especificado
    
    Args:
        data_inicio: Data de início da reserva
        data_fim: Data de fim da reserva
        permitir_dia_devolucao: Se True, permite reservas no mesmo dia da devolução anterior
    """
    if permitir_dia_devolucao:
        query = """
            SELECT c.id, c.marca, c.modelo, c.placa, c.diaria, c.preco_km, c.cor
            FROM carros c
            WHERE c.status NOT IN ('Indisponível', 'Excluído')
            AND c.id NOT IN (
                SELECT r.carro_id FROM reservas r
                WHERE r.reserva_status IN ('Reservada', 'Locada')
                AND (
                    -- Lógica que permite reserva no mesmo dia da devolução
                    (r.data_inicio < %s)
                    AND 
                    (r.data_fim > %s)
                )
            )
            ORDER BY c.marca, c.modelo
        """
    else:
        # Comportamento original
        query = """
            SELECT c.id, c.marca, c.modelo, c.placa, c.diaria, c.preco_km, c.cor
            FROM carros c
            WHERE c.status NOT IN ('Indisponível', 'Excluído')
            AND c.id NOT IN (
                SELECT r.carro_id FROM reservas r
                WHERE r.reserva_status IN ('Reservada', 'Locada')
                AND (r.data_inicio <= %s AND r.data_fim >= %s)
            )
            ORDER BY c.marca, c.modelo
        """
    
    return run_query_dataframe(query, (data_fim, data_inicio))

def format_vehicle_options(df_veiculos):
    """
    Formata as opções de veículos para exibição no selectbox
    """
    if df_veiculos.empty:
        return ["Nenhum veículo disponível"]
    
    return df_veiculos.apply(
        lambda r: f"{r['id']} - {r['marca']} {r['modelo']} ({r['placa']}) – {formatar_moeda(r['diaria'])}/dia",
        axis=1
    ).tolist()

def check_vehicle_availability(carro_id, data_inicio, data_fim, reserva_id_to_exclude=None):
    """
    Verifica se um veículo está disponível para o período
    """
    query = """
        SELECT COUNT(*) as count
        FROM reservas 
        WHERE carro_id = %s 
        AND reserva_status IN ('Reservada', 'Locada')
        AND (data_inicio <= %s AND data_fim >= %s)
    """
    
    params = [carro_id, data_fim, data_inicio]
    
    if reserva_id_to_exclude:
        query += " AND id != %s"
        params.append(reserva_id_to_exclude)
    
    result = run_query_dataframe(query, params)
    
    return result.iloc[0]['count'] == 0 if not result.empty else True


def gerar_recibo_para_download(reserva_id):
    # 1. Buscar dados da reserva
    query_reserva = """
        SELECT 
            r.id, r.data_inicio, r.data_fim, r.km_saida, r.km_volta, r.km_franquia,
            r.custo_lavagem, r.valor_multas, r.valor_danos, r.valor_outros, r.adiantamento, r.valor_total,
            cl.nome AS cliente_nome, cl.cpf AS cliente_cpf, cl.telefone AS cliente_telefone,
            c.marca AS carro_marca, c.modelo AS carro_modelo, c.placa AS carro_placa, c.cor AS carro_cor, c.preco_km AS carro_preco_km, c.diaria AS carro_diaria,
            c.numero_chassi AS carro_chassi, c.numero_renavam AS carro_renavam
        FROM reservas r
        JOIN clientes cl ON r.cliente_id = cl.id
        JOIN carros c ON r.carro_id = c.id
        WHERE r.id = %s
    """
    reserva_df = run_query_dataframe(query_reserva, (reserva_id,))

    if reserva_df.empty:
        st.error(f"Reserva com ID {reserva_id} não encontrada.")
        return None

    reserva_data = reserva_df.iloc[0].to_dict()

    # Preparar dados do cliente
    cliente = {
        'nome': reserva_data['cliente_nome'],
        'cpf': reserva_data['cliente_cpf'],
        'telefone': reserva_data['cliente_telefone']
    }

    # Preparar dados do carro
    carro = {
        'marca': reserva_data['carro_marca'],
        'modelo': reserva_data['carro_modelo'],
        'placa': reserva_data['carro_placa'],
        'cor': reserva_data['carro_cor'],
        'preco_km': reserva_data['carro_preco_km'],
        'diaria': reserva_data['carro_diaria'],
        'chassi': reserva_data['carro_chassi'],
        'renavam': reserva_data['carro_renavam']
    }

    # Calcular dias de cobrança
    data_inicio_dt = pd.to_datetime(reserva_data['data_inicio']).date()
    data_fim_dt = pd.to_datetime(reserva_data['data_fim']).date()
    dias_cobranca = (data_fim_dt - data_inicio_dt).days

    # Preparar dados para o recibo
    recibo_dados = {
        'data_inicio': data_inicio_dt,
        'data_fim': data_fim_dt,
        'km_saida': reserva_data['km_saida'],
        'km_volta': reserva_data['km_volta'],
        'km_franquia': reserva_data['km_franquia'] if reserva_data['km_franquia'] is not None else 0.0,
        'dias_cobranca': dias_cobranca,
        'custo_diarias': reserva_data['carro_diaria'] * dias_cobranca,
        'custo_km': max(0, reserva_data['km_volta'] - reserva_data['km_saida'] - (reserva_data['km_franquia'] if reserva_data['km_franquia'] is not None else 0.0)) * reserva_data['carro_preco_km'],
        'valor_lavagem': reserva_data['custo_lavagem'] if reserva_data['custo_lavagem'] is not None else 0.0,
        'valor_multas': reserva_data['valor_multas'] if reserva_data['valor_multas'] is not None else 0.0,
        'valor_danos': reserva_data['valor_danos'] if reserva_data['valor_danos'] is not None else 0.0,
        'valor_outros': reserva_data['valor_outros'] if reserva_data['valor_outros'] is not None else 0.0,
        'adiantamento': reserva_data['adiantamento'] if reserva_data['adiantamento'] is not None else 0.0,
        'total_final': reserva_data['valor_total']
    }

    # Gerar o PDF
    try:
        pdf_bytes = gerar_recibo_pdf(cliente, carro, recibo_dados)
        return pdf_bytes
    except Exception as e:
        st.error(f"Erro ao gerar recibo PDF: {e}")
        return None

with st.sidebar:
    st.markdown(
        """
        <div class="sidebar-logo">
            <div class="sidebar-logo__title">Locadora Iguacu</div>
            <div class="sidebar-logo__subtitle">Operação e Gestão</div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    menu_options = [
        "Dashboard",
        "Clientes",
        "Frota (Carros)",
        "Reservas",
        "Entrega do veículo",
        "Devolução",
        "Histórico",
        "Relatórios",
        "Gerenciar Multas"
    ]
    if current_user.get('role') == 'admin':
        menu_options.append("Gerenciar Usuários")
    if 'menu_initialized' not in st.session_state:
        st.session_state.menu_initialized = True
        st.session_state.main_menu_selector = "Dashboard"
    menu = st.radio("Navegação principal", menu_options, key="main_menu_selector", index=0)
    st.markdown("---")
    st.markdown("#### Usuário")
    st.markdown(f"**{current_user['full_name']}**")
    st.caption(f"Perfil: {current_user['role'].title()}")
    if st.button("🚪 Sair do Sistema", key="logout_btn", width='stretch'):
        logout_user()

# Funções utilitárias globais
def validar_cnh_simplificada(dados_cliente):
    validade = dados_cliente.get('validade_cnh')
    if not validade:
        st.warning("Validade da CNH não informada.")
        return False
    validade_date = pd.to_datetime(validade).date()
    if validade_date < date.today():
        st.error(f"CNH expirada em {validade_date.strftime('%d/%m/%Y')}.")
        return False
    return True

# 0. GERENCIAR USUÁRIOS (APENAS ADMIN)
if menu == "Gerenciar Usuários" and current_user.get('role') == 'admin':
    st.title("👥 Gerenciamento de Usuários")
    
    # Abas para diferentes funcionalidades
    tab_listar, tab_adicionar, tab_logs = st.tabs(["Listar Usuários", "Adicionar Usuário", "Logs de Atividade"])
    
    with tab_listar:
        st.subheader("Lista de Usuários")
        
        # Buscar todos os usuários
        users = supabase_auth_manager.get_users()
        
        if not users:
            st.info("Nenhum usuário cadastrado.")
        else:
            # Converter para DataFrame para exibição
            df_usuarios = pd.DataFrame(users)
            df_usuarios['Criado em'] = pd.to_datetime(df_usuarios['created_at']).dt.strftime('%d/%m/%Y %H:%M')
            df_usuarios['Último Login'] = pd.to_datetime(df_usuarios['last_login']).dt.strftime('%d/%m/%Y %H:%M')
            
            # Renomear colunas para exibição
            df_usuarios = df_usuarios.rename(columns={
                'id': 'ID',
                'full_name': 'Nome Completo',
                'email': 'E-mail',
                'role': 'Função',
                'created_at': 'Data de Criação',
                'last_login': 'Último Acesso'
            })
            
            # Adicionar coluna de status ativo (todos os usuários retornados estão ativos)
            df_usuarios['Status'] = 'Ativo'
            
            # Selecionar e ordenar colunas para exibição
            colunas_exibicao = ['ID', 'Nome Completo', 'E-mail', 'Função', 'Status', 'Data de Criação', 'Último Acesso']
            
            # Exibir tabela de usuários
            st.dataframe(
                df_usuarios[colunas_exibicao],
                width='stretch',
                hide_index=True
            )
            
            # Nota: A edição de usuários deve ser feita diretamente no Supabase
            st.info("ℹ️ A edição de usuários (nome, e-mail, função, etc.) deve ser realizada diretamente no painel do Supabase para garantir a integridade dos dados.")

    with tab_logs:
        st.subheader("📋 Logs de Atividade do Sistema")
        st.write("Registro de todas as ações realizadas no sistema.")
        
        # Filtros
        with st.expander("🔍 Filtros", expanded=True):
            col1, col2, col3 = st.columns(3)
            
            # Filtro por usuário
            all_users = supabase_auth_manager.get_users()
            user_options = ["Todos"] + [f"{u['email']} ({u.get('full_name', 'Sem nome')})" for u in all_users if 'email' in u]
            selected_user = col1.selectbox("Filtrar por usuário", user_options)

            # Get the selected user's email
            selected_email = selected_user.split(" (")[0] if selected_user != "Todos" else None
            user_filter = f"AND u.email = '{selected_email}'" if selected_email else ""
            
            # Filtro por status de sessão
            session_status = ["Todas", "Ativas", "Expiradas"]
            selected_status = col2.selectbox("Filtrar por status", session_status)
            status_filter = "AND s.last_activity > NOW() - INTERVAL '24 hours'" if selected_status == "Ativas" else ("AND s.last_activity <= NOW() - INTERVAL '24 hours'" if selected_status == "Expiradas" else "")
            
            # Filtro por data
            today = datetime.now().date()
            date_range = col3.date_input(
                "Período",
                value=(today - timedelta(days=7), today),
                format="DD/MM/YYYY"
            )
        
        # Construir query de filtro
        user_filter = f"AND u.email = '{selected_user.split(' (')[0]}'" if selected_user != "Todos" else ""
        
        # Verificar se temos um intervalo de datas válido
        if isinstance(date_range, tuple) and len(date_range) == 2:
            start_date, end_date = date_range
            date_filter = f"AND s.created_at BETWEEN '{start_date}' AND '{end_date} 23:59:59'"
        else:
            date_filter = f"AND s.created_at >= '{today - timedelta(days=7)}'"
        
        # Buscar logs de sessões com filtros aplicados
        logs_query = f"""
            SELECT 
                s.created_at as timestamp,
                u.email,
                u.full_name,
                'session_' || split_part(s.id::text, '-', 1) as session_id,
                s.ip_address,
                s.device_info,
                s.last_activity
            FROM public.sessions s
            JOIN public.users u ON s.user_id = u.id
            WHERE 1=1
            {user_filter}
            {status_filter}
            {date_filter}
            ORDER BY s.created_at DESC
            LIMIT 1000
        """
        
        logs = run_query(logs_query, fetch=True)
        
        if isinstance(logs, str):
            st.error(f"Erro ao buscar logs: {logs}")
            logs_display = pd.DataFrame()
        elif not logs.empty:
            # Formatar dados para exibição
            logs_display = logs.copy()
            logs_display['timestamp'] = pd.to_datetime(logs_display['timestamp']).dt.strftime('%d/%m/%Y %H:%M:%S')
            logs_display['last_activity'] = pd.to_datetime(logs_display['last_activity']).dt.strftime('%d/%m/%Y %H:%M:%S')
            
            # Exibir tabela de logs
            st.dataframe(
                logs_display[['timestamp', 'email', 'full_name', 'session_id', 'ip_address', 'last_activity']],
                column_config={
                    'timestamp': 'Login',
                    'email': 'E-mail',
                    'full_name': 'Nome',
                    'session_id': 'Sessão',
                    'ip_address': 'Endereço IP',
                    'last_activity': 'Última Atividade'
                },
                width='stretch',
                hide_index=True
            )
            
            # Estatísticas rápidas
            st.subheader("📊 Estatísticas")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                total_sessions = len(logs)
                st.metric("Total de Sessões", total_sessions)
                
            with col2:
                unique_users = logs['email'].nunique()
                st.metric("Usuários Únicos", unique_users)
                
            with col3:
                if 'ip_address' in logs.columns and not logs['ip_address'].isna().all():
                    unique_ips = logs['ip_address'].nunique()
                    st.metric("IPs Únicos", unique_ips)
                else:
                    active_sessions = len(logs[logs['last_activity'] > (datetime.now() - timedelta(hours=24))])
                    st.metric("Sessões Ativas", active_sessions)
                
        else:
            st.info("Nenhum registro de log encontrado com os filtros ativos.")
    
    with tab_adicionar:
        st.subheader("Adicionar Novo Usuário")
        
        with st.form("form_novo_usuario"):
            novo_username = st.text_input("Nome de Usuário*")
            novo_nome = st.text_input("Nome Completo*")
            novo_email = st.text_input("E-mail")
            nova_senha = st.text_input("Senha*", type="password")
            confirmar_senha = st.text_input("Confirmar Senha*", type="password")
            # Opções de função (excluindo admin como opção para novos usuários)
            role_options = list(USER_ROLES.items())[1:]
            
            nova_funcao = st.selectbox(
                "Função*",
                role_options,
                format_func=lambda x: x[1]  # Mostrar o nome amigável da função
            )
            nova_funcao = nova_funcao[0]  # Manter apenas a chave da função
            
            if st.form_submit_button("Criar Usuário", type="primary"):
                # Validações
                if not all([novo_username, novo_nome, nova_senha, confirmar_senha]):
                    st.error("Por favor, preencha todos os campos obrigatórios (*).")
                elif nova_senha != confirmar_senha:
                    st.error("As senhas não conferem.")
                elif len(nova_senha) < 6:
                    st.error("A senha deve ter pelo menos 6 caracteres.")
                else:
                    # Criar usuário
                    sucesso, mensagem = supabase_auth_manager.create_user(
                        username=novo_username,
                        password=nova_senha,
                        role=nova_funcao,
                        full_name=novo_nome,
                        email=novo_email
                    )
                    
                    if sucesso:
                        st.success(mensagem)
                        st.rerun()  # Recarregar a página para atualizar a lista
                    else:
                        st.error(f"Erro ao criar usuário: {mensagem}")

# 1. DASHBOARD
elif menu == "Dashboard":
    st.title("📊 Painel Gerencial e Agenda do Dia")

    # 1. Métricas Principais - Otimizadas com única query
    dashboard_data = get_dashboard_data()
    
    # Para as listas detalhadas (agenda), ainda precisamos das consultas completas
    hoje_str = date.today().strftime('%Y-%m-%d')

    # Busca Locados (reserva_status = 'Locada')
    query_locados = f"""
        SELECT
            CONCAT(c.marca, ' ', c.modelo) as modelo, c.placa, r.data_fim, cl.nome as cliente
        FROM reservas r
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status = 'Ativa' AND r.reserva_status = 'Locada'
    """
    df_locados = run_query(query_locados, fetch=True)
    df_locados = df_locados if not isinstance(df_locados, str) else pd.DataFrame()

    # Busca Reservados (reserva_status = 'Reservada')
    query_reservados = f"""
        SELECT
            CONCAT(c.marca, ' ', c.modelo) as modelo, c.placa, r.data_inicio, cl.nome as cliente
        FROM reservas r
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status = 'Ativa' AND r.reserva_status = 'Reservada'
    """
    df_reservados = run_query(query_reservados, fetch=True)
    df_reservados = df_reservados if not isinstance(df_reservados, str) else pd.DataFrame()

    # Consultas de Agenda (usando o novo reserva_status)
    # Entradas Previstas são reservas que estão sendo devolvidas hoje
    query_entradas_hoje = """
        SELECT c.modelo, c.placa, cl.nome as cliente, r.data_inicio, r.data_fim, r.reserva_status
        FROM carros c
        JOIN reservas r ON c.id = r.carro_id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.status = 'Ativa' AND r.reserva_status = 'Locada' AND r.data_fim::date = %s::date
    """
    entradas_hoje = run_query(query_entradas_hoje, (hoje_str,), fetch=True)

    # Saídas Previstas são reservas que precisam ser entregues hoje
    query_saidas_hoje = """
        SELECT c.modelo, c.placa, cl.nome as cliente, r.data_inicio, r.data_fim, r.reserva_status
        FROM carros c
        JOIN reservas r ON c.id = r.carro_id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE r.reserva_status = 'Reservada' AND r.data_inicio::date = %s::date
    """
    saidas_hoje = run_query(query_saidas_hoje, (hoje_str,), fetch=True)

    df_entradas = entradas_hoje if not isinstance(entradas_hoje, str) else pd.DataFrame()
    df_saidas = saidas_hoje if not isinstance(saidas_hoje, str) else pd.DataFrame()

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Veículos na Frota", dashboard_data['total_carros'])
    col2.metric("Carros Locados Agora", dashboard_data['carros_locados'])
    col3.metric("Carros Reservados", dashboard_data['carros_reservados'])
    col4.metric(f"Faturamento {datetime.now().strftime('%b/%Y')}", formatar_moeda(dashboard_data['faturamento_mensal']))
    col5.metric("Devoluções Previstas Hoje", dashboard_data['devolucoes_hoje'])

    st.divider()

    # --- CHECAGEM RÁPIDA DE DISPONIBILIDADE (CORRIGIDO) ---
    st.subheader("🗓️ Verificação Rápida de Disponibilidade")

    col_data1, col_data2 = st.columns(2)

    data_inicio_check = col_data1.date_input("Início da Locação", date.today(), key="check_inicio")
    data_fim_check = col_data2.date_input("Fim da Locação", data_inicio_check + timedelta(days=1),
                                          min_value=data_inicio_check, key="check_fim")
    #data_fim_check = col_data2.date_input("Fim da Locação", date.today() + timedelta(days=7),
                                          #min_value=data_inicio_check, key="check_fim")

    # Adiciona o toggle para permitir reserva no mesmo dia da devolução
    st.markdown("---")
    with st.expander("⚙️ Opções de Disponibilidade", expanded=False):
        st.caption("Configurações avançadas de disponibilidade de veículos")
        permitir_dia_devolucao = st.toggle(
            "🔧 Permitir reserva no mesmo dia da devolução",
            value=False,
            help="Quando ativado, permite que um veículo seja reservado no mesmo dia em que retorna de outra locação. Use com cautela."
        )
        if permitir_dia_devolucao:
            st.warning("""
            ⚠️ **Modo experimental ativado**
            
            Esta opção permite que veículos sejam reservados no mesmo dia em que retornam de outra locação. 
            
            **Atenção:** Verifique cuidadosamente as reservas existentes para evitar sobreposições indesejadas.
            """)

    if data_inicio_check <= data_fim_check:
        # Usa a função get_available_vehicles que já tem a lógica de verificação
        carros_disponiveis = get_available_vehicles(
            data_inicio_check, 
            data_fim_check,
            permitir_dia_devolucao=permitir_dia_devolucao
        )
        
        # Converte o resultado para o formato esperado pelo código existente
        if not carros_disponiveis.empty:
            # Cria um DataFrame para exibição
            df_livres = carros_disponiveis.copy()
            
            # Formata as colunas de preço
            df_livres['Diária (R$)'] = df_livres['diaria'].apply(lambda x: formatar_moeda(x).replace('R$ ', ''))
            df_livres['Preço/KM (R$)'] = df_livres['preco_km'].apply(lambda x: formatar_moeda(x).replace('R$ ', ''))
            
            st.success(f"✅ {len(df_livres)} Veículos Disponíveis de {data_inicio_check.strftime('%d/%m')} a {data_fim_check.strftime('%d/%m')}.")
            
            # Exibe a tabela formatada
            st.dataframe(
                df_livres[['modelo', 'placa', 'Diária (R$)', 'Preço/KM (R$)']]
                .rename(columns={
                    'modelo': 'Modelo',
                    'placa': 'Placa'
                }),
                width='stretch',
                hide_index=True
            )
        else:
            st.warning("⚠️ Nenhum veículo disponível para o período selecionado.")

    st.divider()

    # 2. Situação de Carros Locados e Reservados
    col_status1, col_status2 = st.columns(2)

    with col_status1:
        st.subheader("Situação da Frota: Carros Locados")
        if not df_locados.empty:
            df_locados['data_fim'] = pd.to_datetime(df_locados['data_fim']).dt.strftime('%d/%m/%Y')

            st.dataframe(
                df_locados.rename(columns={
                    'modelo': 'Modelo',
                    'placa': 'Placa',
                    'cliente': 'Cliente',
                    'data_fim': 'Devolução Prevista'
                }),
                width='stretch'
            )
        else:
            st.success("Nenhum veículo locado no momento.")

    with col_status2:
        st.subheader("Situação da Frota: Carros Reservados (Aguardando Entrega)")
        if not df_reservados.empty:
            df_reservados['data_inicio'] = pd.to_datetime(df_reservados['data_inicio']).dt.strftime('%d/%m/%Y')

            st.dataframe(
                df_reservados.rename(columns={
                    'modelo': 'Modelo',
                    'placa': 'Placa',
                    'cliente': 'Cliente',
                    'data_inicio': 'Data Prevista da Entrega'
                }),
                width='stretch'
            )
        else:
            st.info("Nenhuma reserva pendente de entrega.")

    st.divider()

    # 3. Agenda do Dia
    st.subheader("Agenda: Entradas e Saídas do Dia (HOJE)")
    col_agenda1, col_agenda2 = st.columns(2)

    with col_agenda1:
        st.markdown("##### 📥 Devoluções Previstas (HOJE)")
        if not df_entradas.empty:
            st.dataframe(df_entradas.rename(columns={'modelo': 'Modelo', 'placa': 'Placa', 'cliente': 'Cliente'}),
                         width='stretch')
        else:
            st.info("Nenhuma devolução agendada.")

    with col_agenda2:
        st.markdown("##### 📤 Entregas Agendadas (HOJE)")
        if not df_saidas.empty:
            st.dataframe(df_saidas.rename(columns={'modelo': 'Modelo', 'placa': 'Placa', 'cliente': 'Cliente'}),
                         width='stretch')
        else:
            st.info("Nenhuma nova locação agendada.")


# 2. CLIENTES (COMPLETO: CADASTRO, EDIÇÃO, EXCLUSÃO E OBSERVAÇÕES)
elif menu == "Clientes":
    st.title("👥 Gestão de Clientes")


    tab1, tab2 = st.tabs(["Cadastrar Novo", "Ver / Editar Clientes"])

    with tab1:
        st.subheader("Cadastro de Cliente")
        with st.form("form_cliente"):
            c1, c2 = st.columns(2)
            nome = c1.text_input("Nome Completo*")
            cpf = c2.text_input("CPF* (apenas números)", max_chars=11)
            
            c3, c4 = st.columns(2)
            rg = c3.text_input("RG (apenas números)", max_chars=20)
            cnh = c4.text_input("Número da CNH*")
            
            c5, c6 = st.columns(2)
            validade_cnh = c5.date_input("Validade da CNH*", min_value=date.today())
            uf_cnh = c6.selectbox("UF da CNH*", ["", "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"])
            
            telefone = st.text_input("Telefone* (com DDD)")
            endereco = st.text_area("Endereço Completo")
            observacoes = st.text_area("Observações sobre o Cliente (Não obrigatório)", help="NAO APARECEU, CANCELOU A RESERVA, NAO PAGOU A LOCACAO, NAO PAGOU A MULTA, ETC.")
            
            submit = st.form_submit_button("Salvar Cliente", type="primary")

            if submit:
                # Validação de campos obrigatórios
                if not nome or not cpf or not cnh or not telefone:
                    st.error("⚠️ Os campos Nome, CPF, CNH e Telefone são obrigatórios!")
                else:
                    # Verifica se o CPF já existe
                    cpf_existente = run_query(
                        "SELECT 1 FROM clientes WHERE cpf = %s AND status != %s", 
                        (cpf, 'REMOVIDO'), 
                        fetch=True
                    )
                    
                    # Verifica se o RG já existe (se informado)
                    if rg:
                        rg_existente = run_query(
                            "SELECT 1 FROM clientes WHERE rg = %s AND status != %s AND rg IS NOT NULL",
                            (rg, 'REMOVIDO'),
                            fetch=True
                        )
                        if not rg_existente.empty:
                            st.error("⚠️ Já existe um cliente ativo com este RG cadastrado.")
                            st.stop()

                    if not cpf_existente.empty:
                        st.error("❌ Este CPF já está cadastrado no sistema.")
                    else:
                        # Incluindo todos os campos na query INSERT
                        if not uf_cnh:
                            st.error("⚠️ O campo UF da CNH é obrigatório.")
                            st.stop()
                            
                        res = run_query(
                            """
                            INSERT INTO clientes 
                            (nome, cpf, rg, cnh, validade_cnh, uf_cnh, telefone, endereco, observacoes) 
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                            """,
                            (nome, cpf, rg if rg else None, cnh, validade_cnh, uf_cnh, telefone, endereco, observacoes)
                        )
                        if isinstance(res, str):
                            st.error(f"Erro ao cadastrar cliente. Detalhe: {res}")
                        else:
                            st.toast(f"Cliente {nome} cadastrado com sucesso!", icon="✅")
                            st.success(f"Cliente {nome} cadastrado com sucesso!")
                            time.sleep(1)
                            st.rerun()

    with tab2:
        # Filtra para não exibir clientes com status 'Removido' na lista principal
        df_clientes = run_query_dataframe("SELECT * FROM clientes WHERE status != %s", (STATUS_CLIENTE['REMOVIDO'],))
        if not df_clientes.empty:

            # Formatação para exibição
            df_clientes_display = df_clientes.copy()
            df_clientes_display['validade_cnh'] = pd.to_datetime(df_clientes_display['validade_cnh']).dt.strftime(
                '%d/%m/%Y')

            st.dataframe(df_clientes_display, width='stretch')

            cliente_opcoes = df_clientes['id'].astype(str) + " - " + df_clientes['nome']
            opcoes_com_placeholder = ["Selecione o cliente..."] + cliente_opcoes.tolist()

            cliente_sel = st.selectbox("Selecione para Edição ou Exclusão", opcoes_com_placeholder)

            if cliente_sel != "Selecione o cliente...":
                id_cliente_sel = int(cliente_sel.split(" - ")[0])
                dados_atuais = df_clientes[df_clientes['id'] == id_cliente_sel].iloc[0]

                # --- FORMULÁRIO DE EDIÇÃO ---
                st.markdown("---")
                st.subheader(f"✏️ Editando: {dados_atuais['nome']}")

                # Prepara o valor da data da CNH
                if dados_atuais['validade_cnh']:
                    validade_cnh_atual = pd.to_datetime(dados_atuais['validade_cnh']).date()
                else:
                    validade_cnh_atual = date.today()

                with st.form("form_edit_cliente"):
                    # Colunas para campos que não devem ser alterados (CPF) ou são chave (CNH)
                    e1, e2 = st.columns(2)
                    up_nome = e1.text_input("Nome Completo", value=dados_atuais['nome'])
                    e2.text_input("CPF (Não Editável)", value=dados_atuais['cpf'], disabled=True)

                    e3, e4 = st.columns(2)
                    up_rg = e3.text_input("RG (apenas números)", value=dados_atuais.get('rg', ''), max_chars=20)
                    up_cnh = e4.text_input("Número da CNH", value=dados_atuais['cnh'])
                    
                    e5, e6, e7 = st.columns(3)
                    up_validade_cnh = e5.date_input("Validade CNH", value=validade_cnh_atual)
                    if up_validade_cnh < date.today():
                        e5.warning("Atenção: A validade da CNH está vencida!")
                    
                    # Obtém o valor atual da UF da CNH ou define como vazio se não existir
                    uf_cnh_atual = dados_atuais.get('uf_cnh', '')
                    up_uf_cnh = e6.selectbox(
                        "UF da CNH*", 
                        ["", "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"],
                        index=["", "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO", "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI", "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO"].index(uf_cnh_atual) if uf_cnh_atual else 0
                    )
                    
                    up_telefone = e7.text_input("Telefone* (com DDD)", value=dados_atuais['telefone'])

                    up_endereco = st.text_area("Endereço Completo", value=dados_atuais.get('endereco', ''))

                    # Carrega o valor atual para observações (usando .get() para evitar KeyError)
                    up_observacoes = st.text_area("Observações", value=dados_atuais.get('observacoes', ''), help="NAO APARECEU, CANCELOU A RESERVA, NAO PAGOU A LOCACAO, NAO PAGOU A MULTA, ETC.")

                    col_botoes = st.columns(2)

                    if col_botoes[0].form_submit_button("🔄 Atualizar Dados do Cliente", type="primary"):
                        if not up_nome or not up_cnh or not up_telefone:
                            st.error("⚠️ Os campos Nome, CNH e Telefone não podem ficar vazios.")
                        else:
                            if not up_uf_cnh:
                                st.error("⚠️ O campo UF da CNH é obrigatório.")
                            else:
                                run_query("""
                                    UPDATE clientes
                                    SET nome=%s, 
                                        rg=%s, 
                                        cnh=%s, 
                                        validade_cnh=%s, 
                                        uf_cnh=%s,
                                        telefone=%s, 
                                        endereco=%s, 
                                        observacoes=%s
                                    WHERE id=%s
                                """, (
                                    up_nome, 
                                    up_rg if up_rg else None, 
                                    up_cnh, 
                                    up_validade_cnh, 
                                    up_uf_cnh,
                                    up_telefone, 
                                    up_endereco, 
                                    up_observacoes, 
                                    id_cliente_sel
                                ))
                            st.toast("Cliente atualizado!", icon="✔️")
                            st.success(f"Cliente **{up_nome}** atualizado com sucesso!")
                            st.rerun()

                    if col_botoes[1].form_submit_button("🗑️ Marcar como REMOVIDO"):
                        # --- CHECAGEM CRÍTICA DE RESERVAS ATIVAS (REPETIDA DA LÓGICA ANTERIOR) ---
                        reservas_ativas_check = run_query(
                            "SELECT COUNT(*) FROM reservas WHERE cliente_id=%s AND reserva_status IN ('Reservada', 'Locada')",
                            (id_cliente_sel,),
                            fetch=True
                        ).iloc[0, 0]

                        if reservas_ativas_check > 0:
                            st.error(
                                f"❌ Não é possível remover. O cliente possui {reservas_ativas_check} reserva(s) Ativa(s). Finalize a devolução primeiro.")
                        else:
                            # Se não há reservas ativas, marca o cliente como 'Removido'
                            run_query("UPDATE clientes SET status=%s WHERE id=%s", (STATUS_CLIENTE['REMOVIDO'], id_cliente_sel))
                            st.toast("Cliente marcado como Removido!", icon="🗑️")
                            st.warning("Cliente marcado como **REMOVIDO** (Registro mantido para histórico).")
                            st.rerun()
        else:
            st.info("Nenhum cliente cadastrado ou ativo.")


# 3. FROTA (CARROS) - COMPLETO COM STATUS 'EXCLUÍDO'
elif menu == "Frota (Carros)":
    st.title("🚙 Gestão da Frota")

    tab1, tab2 = st.tabs(["Cadastrar Veículo", "Ver / Editar / Status"])

    with tab1:
        st.subheader("Cadastro de Novo Veículo")
        with st.form("cadastro_carro", clear_on_submit=False):
            col_a, col_b = st.columns(2)
            marca = col_a.text_input("Marca (ex: Fiat, Volkswagen, Toyota)", help="Informe a marca do veículo conforme consta no documento")
            modelo = col_b.text_input("Modelo (ex: Mobi, Gol, Corolla)", help="Informe o modelo específico do veículo")

            placa = st.text_input("Placa", help="Informe a placa no formato AAA-0000 ou AAA0A00").upper()
            st.caption("ℹ️ A placa será automaticamente convertida para maiúsculas.")

            col_c, col_d = st.columns(2)
            cor = col_c.text_input("Cor", help="Informe a cor predominante do veículo")
            km = col_d.number_input("KM Atual", 0, help="Quilometragem atual do veículo.")
            
            col_e, col_f, col_g = st.columns(3)
            numero_chassi = col_e.text_input("Número do Chassi", help="Número de identificação único do veículo (17 caracteres alfanuméricos)")
            numero_renavam = col_f.text_input("Número do Renavam", help="Número de identificação do veículo no DENATRAN (11 dígitos)")
            km_troca_oleo = col_g.number_input("KM da Próxima Troca de Óleo", min_value=0, value=0, 
                                             help="Próxima troca de óleo baseada na quilometragem")

            ano_veiculo = st.number_input("Ano do Veiculo", min_value=1900, max_value=date.today().year + 1, 
                                        value=date.today().year, help="Ano de fabricação do veículo")

            col_h, col_i = st.columns(2)
            diaria = col_h.number_input("Valor Diária (R$)", 0.0, help="Valor cobrado por dia de locação (R$)")
            p_km = col_i.number_input("Custo por KM (R$)", 0.0, help="Valor cobrado por quilômetro rodado (R$)")
            st.caption("💡 Dica: Considere os custos de manutenção e depreciação ao definir os valores de locação.")

            # Define o status inicial no cadastro
            status_help = """
            - **Disponível**: Veículo pronto para locação
            - **Em Manutenção**: Em manutenção ou revisão
            - **Indisponível**: Temporariamente fora de circulação
            - **Reservado**: Com reserva ativa
            - **Locado**: Atualmente alugado
            """
            status_inicial = st.selectbox("Status Inicial (Padrão)", options=list(STATUS_CARRO.values()), 
                                       index=0, help=status_help)

            if st.form_submit_button("Salvar Carro", type="primary"):
                if not marca or not modelo or not placa or not cor or diaria <= 0 or not numero_chassi or not numero_renavam or not ano_veiculo:
                    st.error("⚠️ Preencha Marca, Modelo, Placa, Cor, Número do Chassi, Número do Renavam, Ano do Veiculo e certifique-se que a Diária seja maior que zero.")
                else:
                    res = run_query(
                        "INSERT INTO carros (marca, modelo, placa, cor, km_atual, diaria, preco_km, status, numero_chassi, numero_renavam, ano_veiculo, km_troca_oleo) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                        (marca, modelo, placa, cor, km, diaria, p_km, status_inicial, numero_chassi, numero_renavam, ano_veiculo, km_troca_oleo)
                    )
                    if isinstance(res, str):
                        st.error(f"Erro ao cadastrar. Detalhe: {res}")
                    else:
                        st.toast(f"Veículo {marca} {modelo} cadastrado com sucesso!", icon="✅")
                        st.success(f"Veículo {marca} {modelo} cadastrado com sucesso!")
                        time.sleep(1)  # Pequeno atraso para exibir a mensagem
                        st.rerun()

    with tab2:
        df = run_query("SELECT * FROM carros WHERE status != %s", (STATUS_CARRO['EXCLUIDO'],), fetch=True)
        
        if not isinstance(df, str) and not df.empty:
            st.subheader("Frota Atual")
            # Adiciona coluna calculada para KM até próxima troca de óleo
            df_display = df.copy()
            df_display['km_ate_proxima_troca'] = df_display.apply(
                lambda row: max(0, int(row['km_troca_oleo']) - int(row['km_atual'])) if pd.notna(row['km_troca_oleo']) and pd.notna(row['km_atual']) else 0,
                axis=1
            )

            # Exibe apenas as colunas principais incluindo a nova coluna calculada
            st.dataframe(df_display[['id', 'marca', 'modelo', 'placa', 'cor', 'km_atual', 'km_troca_oleo', 'km_ate_proxima_troca', 'diaria', 'status', 'numero_chassi', 'numero_renavam', 'ano_veiculo']], width='stretch')

            carro_opcoes = df['id'].astype(str) + " - " + df['marca'] + " " + df['modelo'] + " (" + df['placa'] + ")"
            opcoes_com_placeholder = ["Selecione o veículo..."] + carro_opcoes.tolist()

            carro_sel = st.selectbox("Selecione Veículo para Ação", opcoes_com_placeholder)

            if carro_sel != "Selecione o veículo...":
                id_edit = int(carro_sel.split(" - ")[0])
                dados_atuais = df[df['id'] == id_edit].iloc[0]

                # --- FORMULÁRIO DE EDIÇÃO DE DADOS E STATUS ---
                st.markdown("---")
                st.subheader(f"✏️ Editando: {dados_atuais['marca']} {dados_atuais['modelo']} ({dados_atuais['placa']})")

                with st.form("form_edit_carro"):
                    # 1. Campos de dados
                    st.markdown("##### Dados do Veículo")
                    col_info1, col_info2, col_info3 = st.columns(3)
                    up_marca = col_info1.text_input("Marca", value=dados_atuais['marca'] if dados_atuais['marca'] is not None else "", key="up_marca")
                    up_modelo = col_info2.text_input("Modelo", value=dados_atuais['modelo'], key="up_modelo")
                    up_cor = col_info3.text_input("Cor", value=dados_atuais['cor'], key="up_cor")

                    up_placa = st.text_input("Placa", value=dados_atuais['placa'], key="up_placa")

                    col_chassi, col_renavam, col_ano = st.columns(3)
                    up_numero_chassi = col_chassi.text_input("Número do Chassi", value=dados_atuais['numero_chassi'], key="up_numero_chassi")
                    up_numero_renavam = col_renavam.text_input("Número do Renavam", value=dados_atuais['numero_renavam'], key="up_numero_renavam")
                    up_km_troca_oleo = col_ano.number_input("KM da Próxima Troca de Óleo", min_value=0, value=int(dados_atuais['km_troca_oleo']) if dados_atuais['km_troca_oleo'] is not None and not pd.isna(dados_atuais['km_troca_oleo']) else 10000, key="up_km_troca_oleo")

                    up_ano_veiculo = st.number_input("Ano do Veiculo", min_value=1900, max_value=date.today().year + 1, value=int(dados_atuais['ano_veiculo']) if dados_atuais['ano_veiculo'] is not None and not pd.isna(dados_atuais['ano_veiculo']) else date.today().year, key="up_ano_veiculo")

                    st.markdown("##### Valores e Quilometragem")
                    col_a, col_b = st.columns(2)
                    up_diaria = col_a.number_input("Valor Diária (R$)", value=float(dados_atuais['diaria']),
                                                   key="up_diaria")
                    up_p_km = col_b.number_input("Custo por KM (R$)", value=float(dados_atuais['preco_km']),
                                                 key="up_p_km")
                    up_km = st.number_input("KM Atual", value=int(dados_atuais['km_atual']), key="up_km")

                    st.markdown("---")

                    # 2. Edição de Status
                    st.markdown("##### Status do Veículo")

                    # Encontra o índice do status atual
                    status_index = list(STATUS_CARRO.values()).index(dados_atuais['status']) if dados_atuais[
                                                                                                    'status'] in STATUS_CARRO.values() else 0

                    up_status = st.selectbox(
                        "Alterar Status:",
                        options=list(STATUS_CARRO.values()),
                        index=status_index,
                        key="up_status",
                        help="Define o estado do veículo."
                    )

                    col_botoes = st.columns(2)

                    if col_botoes[0].form_submit_button("🔄 Atualizar Dados e Status", type="primary"):
                        # Verificações de validação
                        if up_km < int(dados_atuais['km_atual']):
                            st.error("❌ O KM atualizado não pode ser menor que o KM registrado anteriormente.")
                        elif not up_marca or not up_modelo or not up_placa:
                            st.error("❌ Marca, Modelo e Placa são obrigatórios.")
                        else:
                            # Verificar se a placa já existe (exceto para o veículo atual)
                            if up_placa != dados_atuais['placa']:
                                placa_check = run_query("SELECT id FROM carros WHERE placa = %s AND id != %s", (up_placa, id_edit), fetch=True)
                                if isinstance(placa_check, str):
                                    st.error(f"Erro ao verificar placa: {placa_check}")
                                elif not placa_check.empty:
                                    st.error(f"❌ Placa {up_placa} já está cadastrada para outro veiculo.")
                                else:
                                    # Atualiza todos os campos
                                    run_query("UPDATE carros SET marca=%s, modelo=%s, placa=%s, cor=%s, diaria=%s, preco_km=%s, km_atual=%s, status=%s, numero_chassi=%s, numero_renavam=%s, ano_veiculo=%s, km_troca_oleo=%s WHERE id=%s",
                                              (up_marca, up_modelo, up_placa, up_cor, up_diaria, up_p_km, up_km, up_status, up_numero_chassi, up_numero_renavam, up_ano_veiculo, up_km_troca_oleo, id_edit))
                                    st.toast("Dados do veiculo atualizados!", icon="✔️")
                                    st.success(f"Veiculo **{up_marca} {up_modelo}** atualizado para status **{up_status}**!")
                                    time.sleep(1)
                                    st.rerun()
                            else:
                                # Atualiza todos os campos
                                run_query("UPDATE carros SET marca=%s, modelo=%s, placa=%s, cor=%s, diaria=%s, preco_km=%s, km_atual=%s, status=%s, numero_chassi=%s, numero_renavam=%s, ano_veiculo=%s, km_troca_oleo=%s WHERE id=%s",
                                          (up_marca, up_modelo, up_placa, up_cor, up_diaria, up_p_km, up_km, up_status, up_numero_chassi, up_numero_renavam, up_ano_veiculo, up_km_troca_oleo, id_edit))
                                st.toast("Dados do veiculo atualizados!", icon="✔️")
                                st.success(f"Veiculo **{up_marca} {up_modelo}** atualizado para status **{up_status}**!")
                                time.sleep(1)
                                st.rerun()

                    # Botão para retirar definitivamente da frota, mudando o status para EXCLUÍDO
                    if col_botoes[1].form_submit_button("🔥 Marcar como EXCLUÍDO (Retirada Definitiva)"):
                        # Verifica se o carro está locado ou reservado
                        if dados_atuais['status'] in [STATUS_CARRO['LOCADO'], STATUS_CARRO['RESERVADO']]:
                            st.error(
                                f"❌ Não é possível excluir. O carro está **{dados_atuais['status']}**. Finalize a pendência primeiro.")
                        else:
                            # Se não está em locação, define o status para 'EXCLUIDO'
                            run_query("UPDATE carros SET status=%s WHERE id=%s",
                                      (STATUS_CARRO['EXCLUÍDO'], id_edit))
                            st.toast("Carro marcado como Excluído!", icon="🔥")
                            st.error("Veículo marcado como **EXCLUÍDO** (Registro mantido para histórico).")
                            time.sleep(1)
                            st.rerun()
        else:
            st.info("Nenhum carro cadastrado na frota.")

elif menu == "Reservas":
    render_section_header(
        title="Reservas",
        subtitle="Cadastre novas reservas e edite locações em andamento.",
        icon="🧾",
        trail=["Reservas"]
    )

    tab_criar, tab_editar = st.tabs(["Reservas Simplificadas", "Gerenciar Reservas"])

    with tab_criar:
        st.markdown("### 📅 Escolha as datas da locação")
        hoje = date.today()
        col_datas = st.columns(2)
        
        inicio = col_datas[0].date_input(
            "Data de retirada",
            value=hoje,
            min_value=hoje,
            key="reserva_simples_inicio"
        )
        fim_default = max(inicio + timedelta(days=1), inicio)
        fim = col_datas[1].date_input(
            "Data de devolução",
            value=fim_default,
            min_value=inicio,
            key="reserva_simples_fim"
        )
        
        # Adiciona opção de permitir reserva no mesmo dia da devolução
        with st.expander("⚙️ Opções de Disponibilidade", expanded=False):
            st.caption("Configurações avançadas de disponibilidade")
            permitir_dia_devolucao = st.toggle(
                "🔧 Permitir reserva no mesmo dia da devolução",
                value=False,
                key="permitir_dia_devolucao_reserva",
                help="Quando ativado, permite que um veículo seja reservado no mesmo dia em que retorna de outra locação. Use com cautela."
            )
            if permitir_dia_devolucao:
                st.warning("""
                ⚠️ **Modo experimental ativado**
                
                Esta opção permite que veículos sejam reservados no mesmo dia em que retornam de outra locação. 
                
                **Atenção:** Verifique cuidadosamente as reservas existentes para evitar sobreposições indesejadas.
                """)

        # Verificar disponibilidade e mostrar veículos disponíveis
        if inicio <= fim:
            carros_disponiveis = get_available_vehicles(
                inicio, 
                fim,
                permitir_dia_devolucao=permitir_dia_devolucao
            )
            
            if not carros_disponiveis.empty:
                st.success(f"✅ {len(carros_disponiveis)} veículos disponíveis para o período selecionado")
                
                # Mostrar cliente selection
                clientes_df = run_query_dataframe("SELECT id, nome, cpf FROM clientes WHERE status = 'Ativo'")
                if clientes_df.empty:
                    st.warning("Cadastre pelo menos um cliente ativo para continuar.")
                else:
                    lista_clientes = ["Selecione o cliente..."] + (
                        clientes_df.apply(lambda c: f"{c['id']} - {c['nome']} (CPF {c['cpf']})", axis=1).tolist()
                    )
                    cliente_sel = st.selectbox("Cliente", lista_clientes, key="reserva_simples_cliente")
                    
                    if cliente_sel != "Selecione o cliente...":
                        cliente_id = int(cliente_sel.split(" - ")[0])
                        
                        # Mostrar veículos disponíveis
                        veiculos_opcoes = format_vehicle_options(carros_disponiveis)
                        carro_sel = st.selectbox("Veículos disponíveis", veiculos_opcoes, key="reserva_simples_carro")
                        
                        if carro_sel != "Nenhum veículo disponível" and carro_sel != "Selecione o veículo...":
                            carro_id = int(carro_sel.split(" - ")[0])
                            dados_carro = carros_disponiveis[carros_disponiveis['id'] == carro_id].iloc[0]
                            
                            st.markdown("### 💰 Detalhes da locação")
                            dias_periodo = max(1, (fim - inicio).days + 0)
                            
                            col_opcoes = st.columns(3)
                            with col_opcoes[0]:
                                meia_diaria = st.checkbox(
                                    "Aplicar meia diária na retirada",
                                    value=False,
                                    key="reserva_simples_meia"
                                )
                            with col_opcoes[1]:
                                desconto = st.number_input(
                                    "Desconto (R$)",
                                    min_value=0.0,
                                    value=0.0,
                                    step=10.0,
                                    format="%.2f",
                                    key="reserva_simples_desconto"
                                )
                            with col_opcoes[2]:
                                km_franquia = st.number_input(
                                    "KM de franquia",
                                    min_value=0,
                                    value=300,
                                    step=50,
                                    key="reserva_simples_km_franquia"
                                )
                            
                            # Cálculo dos valores
                            diaria = Decimal(str(dados_carro['diaria']))
                            
                            # Calculate valor_diarias (dias * valor diaria) - sem desconto
                            if meia_diaria and dias_periodo > 0:
                                valor_diarias = diaria * Decimal(str(dias_periodo - 1)) + (diaria * Decimal('0.5'))
                            else:
                                valor_diarias = diaria * Decimal(str(dias_periodo))
                            
                            # Apply discount to get total_diarias
                            total_diarias = max(Decimal('0'), valor_diarias - Decimal(str(desconto)))
                            
                            # For new reservation, valor_total should equal total_diarias (no additional costs yet)
                            valor_total = total_diarias

                            col_metrics = st.columns(3)
                            col_metrics[0].metric("Diárias", f"{dias_periodo} dia(s)")
                            col_metrics[1].metric("Valor estimado", formatar_moeda(valor_total))
                            col_metrics[2].metric("Franquia KM", f"{km_franquia} km")

                            valor_total_float = float(valor_total)
                            adiantamento_default = float((valor_total * Decimal('0.5'))) if valor_total > 0 else 0.0
                            adiantamento_input = st.number_input(
                                "Adiantamento (R$)",
                                min_value=0.0,
                                max_value=valor_total_float if valor_total_float > 0 else None,
                                value=min(adiantamento_default, valor_total_float) if valor_total_float > 0 else 0.0,
                                step=10.0,
                                format="%.2f",
                                key="reserva_simples_adiantamento"
                            )
                            valor_restante = valor_total - Decimal(str(adiantamento_input))
                            

                            if st.button("Salvar reserva", type="primary", width='stretch', key="btn_reserva_simples"):
                                try:
                                    km_saida_registrado = dados_carro.get('km_atual')
                                    km_saida_registrado = int(km_saida_registrado) if pd.notna(km_saida_registrado) else 0
                                    
                                    nova_reserva_id = run_query(
                                        """
                                        INSERT INTO reservas (
                                            carro_id, cliente_id, data_inicio, data_fim, status, reserva_status,
                                            km_saida, km_franquia, adiantamento,
                                            valor_multas, valor_danos, valor_outros,
                                            desconto_cliente, meia_diaria, total_diarias, valor_total, valor_restante
                                        )
                                        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                                        RETURNING id
                                        """,
                                        (
                                            carro_id,
                                            cliente_id,
                                            inicio,
                                            fim,
                                            'Ativa',
                                            'Reservada',
                                            km_saida_registrado,
                                            km_franquia,
                                            adiantamento_input,
                                            0.0,
                                            0.0,
                                            0.0,
                                            desconto,
                                            meia_diaria,
                                            float(total_diarias),
                                            float(valor_total),
                                            float(valor_restante),
                                        )
                                    )
                                    st.toast("Reserva criada com sucesso!", icon="✅")
                                    st.success(f"Reserva #{nova_reserva_id} confirmada para {inicio.strftime('%d/%m')} → {fim.strftime('%d/%m')}.")
                                    time.sleep(1)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Erro ao salvar reserva: {e}")
            else:
                st.warning("⚠️ Nenhum veículo disponível para o período selecionado. Tente outras datas.")

    with tab_editar:
        st.subheader("Gerenciar reservas em andamento")
        reservas_edicao_df = run_query_dataframe("""
            SELECT 
                r.id, r.carro_id, r.cliente_id, r.data_inicio, r.data_fim, r.reserva_status, r.status,
                r.total_diarias, r.valor_total, r.valor_restante, r.adiantamento,
                r.valor_multas, r.valor_danos, r.valor_outros, r.desconto_cliente, r.km_franquia,
                c.nome AS cliente_nome, c.telefone,
                carros.marca, carros.modelo, carros.placa, carros.diaria
            FROM reservas r
            JOIN clientes c ON r.cliente_id = c.id
            JOIN carros ON r.carro_id = carros.id
            WHERE r.status = 'Ativa'
            ORDER BY r.data_inicio DESC
        """)

        if reservas_edicao_df.empty:
            st.info("Nenhuma reserva ativa encontrada para edição.")
        else:
            df_display = reservas_edicao_df.copy()
            df_display['data_inicio'] = pd.to_datetime(df_display['data_inicio']).dt.strftime('%d/%m/%Y')
            df_display['data_fim'] = pd.to_datetime(df_display['data_fim']).dt.strftime('%d/%m/%Y')
            st.dataframe(
                df_display[[
                    'id', 'cliente_nome', 'modelo', 'placa', 'data_inicio', 'data_fim', 'reserva_status', 'valor_total','adiantamento',  'valor_restante'     
                ]],
                width='stretch'
            )

            opcoes_reserva = ["Selecione uma reserva..."] + reservas_edicao_df.apply(
                lambda r: f"ID {r['id']} • {r['cliente_nome']} ({r['modelo']} - {r['placa']} status: {r['reserva_status']} - {r['data_inicio']} - {r['data_fim']} )",
                axis=1
            ).tolist()
            reserva_escolhida = st.selectbox(
                "Selecione a reserva que deseja editar",
                opcoes_reserva,
                key="gerenciar_reserva_edit"
            )

            if reserva_escolhida != "Selecione uma reserva...":
                try:
                    reserva_id_edicao = int(reserva_escolhida.split(" ")[1])
                except (ValueError, IndexError):
                    st.error("Não foi possível identificar a reserva selecionada.")
                    st.stop()

                dados_reserva_df = run_query_dataframe("SELECT * FROM reservas WHERE id=%s", (reserva_id_edicao,))
                if dados_reserva_df.empty:
                    st.error("Não foi possível carregar os dados completos da reserva.")
                    st.stop()
                dados_reserva = dados_reserva_df.iloc[0].to_dict()

                cliente_atual_id = int(dados_reserva.get('cliente_id'))
                carro_atual_id = int(dados_reserva.get('carro_id'))
                data_inicio_reserva = pd.to_datetime(dados_reserva['data_inicio']).date()
                data_fim_reserva = pd.to_datetime(dados_reserva['data_fim']).date()

                clientes_df = run_query_dataframe("SELECT id, nome FROM clientes ORDER BY nome")
                
                # Inicialmente, mostrar todos os carros disponíveis para as datas atuais da reserva
                carros_query = """
                    SELECT id, marca, modelo, placa, diaria, status
                    FROM carros
                    WHERE status NOT IN (%s, %s)
                    AND id NOT IN (
                        SELECT carro_id FROM reservas
                        WHERE reserva_status IN ('Reservada', 'Locada')
                        AND (data_inicio <= %s AND data_fim >= %s)
                        AND id != %s
                    )
                    ORDER BY marca, modelo
                """
                carros_df = run_query_dataframe(
                    carros_query,
                    (STATUS_CARRO['EXCLUIDO'], STATUS_CARRO['INDISPONIVEL'], data_fim_reserva, data_inicio_reserva, reserva_id_edicao)
                )

                if clientes_df.empty or carros_df.empty:
                    st.error("Cadastre clientes e veículos para editar reservas.")
                    st.stop()

                horario_padrao = dados_reserva.get('horario_entrega')
                if horario_padrao:
                    try:
                        horario_padrao = pd.to_datetime(str(horario_padrao)).time()
                    except Exception:
                        horario_padrao = datetime.strptime("09:00", "%H:%M").time()
                else:
                    horario_padrao = datetime.strptime("09:00", "%H:%M").time()

                status_padrao = dados_reserva.get('status') or 'Ativa'
                reserva_status_padrao = dados_reserva.get('reserva_status') or STATUS_RESERVA['RESERVADA']

                clientes_opcoes = clientes_df.apply(
                    lambda c: f"{c['id']} - {c['nome']}",
                    axis=1
                ).tolist()
                try:
                    idx_cliente = next(i for i, op in enumerate(clientes_opcoes) if op.startswith(f"{cliente_atual_id} -"))
                except StopIteration:
                    idx_cliente = 0

                # Inicializar variáveis que serão definidas dinamicamente após a seleção de datas
                carros_opcoes = []
                idx_carro = 0

                desconto_default = float(dados_reserva.get('desconto_cliente') or 0.0)
                meia_diaria_default = bool(dados_reserva.get('meia_diaria') or False)
                km_franquia_default = int(dados_reserva.get('km_franquia') or 0)
                km_saida_default = int(dados_reserva.get('km_saida') or 0)
                km_volta_default = int(dados_reserva.get('km_volta') or km_saida_default)
                custo_lavagem_default = float(dados_reserva.get('custo_lavagem') or 0.0)
                adiantamento_default = float(dados_reserva.get('adiantamento') or 0.0)
                valor_total_default = float(dados_reserva.get('valor_total') or 0.0)
                valor_multas_default = float(dados_reserva.get('valor_multas') or 0.0)
                valor_danos_default = float(dados_reserva.get('valor_danos') or 0.0)
                valor_outros_default = float(dados_reserva.get('valor_outros') or 0.0)
                valor_diarias_default = float(dados_reserva.get('total_diarias') or 0.0)
                total_diarias_default = float(dados_reserva.get('total_diarias') or 0.0)
                pagamento_parcial_default = float(dados_reserva.get('pagamento_parcial_entrega') or 0.0)
                valor_restante_default = float(dados_reserva.get('valor_restante') or max(0.0, total_diarias_default - adiantamento_default - pagamento_parcial_default))
                status_pagamento_existe = 'status_pagamento' in dados_reserva
                observacoes_existe = 'observacoes' in dados_reserva
                status_pagamento_default = dados_reserva.get('status_pagamento', 'Pendente')
                observacoes_default = dados_reserva.get('observacoes', '') or ''

                st.markdown(f"### ✏️ Editando reserva #{reserva_id_edicao}")

                # Chaves de estado para controlar comportamento reativo de datas
                data_inicio_key = f"data_inicio_edit_{reserva_id_edicao}"
                data_inicio_prev_key = f"{data_inicio_key}_prev"
                data_fim_key = f"data_fim_edit_{reserva_id_edicao}"
                data_fim_prev_key = f"{data_fim_key}_prev"
                data_fim_auto_key = f"{data_fim_key}_auto"
                carro_select_key = f"carro_edit_{reserva_id_edicao}"

                if data_inicio_key not in st.session_state:
                    st.session_state[data_inicio_key] = data_inicio_reserva
                if data_inicio_prev_key not in st.session_state:
                    st.session_state[data_inicio_prev_key] = st.session_state[data_inicio_key]

                if data_fim_key not in st.session_state:
                    st.session_state[data_fim_key] = data_fim_reserva
                if st.session_state[data_fim_key] < st.session_state[data_inicio_key]:
                    st.session_state[data_fim_key] = st.session_state[data_inicio_key] + timedelta(days=1)
                if data_fim_prev_key not in st.session_state:
                    st.session_state[data_fim_prev_key] = st.session_state[data_fim_key]
                if data_fim_auto_key not in st.session_state:
                    st.session_state[data_fim_auto_key] = True

                carro_atual_info_df = carros_df[carros_df['id'] == carro_atual_id]
                if not carro_atual_info_df.empty:
                    carro_atual_label = format_vehicle_options(carro_atual_info_df)[0]
                else:
                    carro_atual_label = f"{carro_atual_id} - Veículo atual"
                if carro_select_key not in st.session_state:
                    st.session_state[carro_select_key] = carro_atual_label

                col_basicos = st.columns(2)
                with col_basicos[0]:
                    cliente_escolhido = st.selectbox(
                        "Cliente",
                        clientes_opcoes,
                        index=idx_cliente,
                        key=f"cliente_edit_{reserva_id_edicao}"
                    )
                    status_opcoes = ["Ativa", "Inativa"]
                    status_escolhido = st.selectbox(
                        "Status operacional",
                        status_opcoes,
                        index=status_opcoes.index(status_padrao) if status_padrao in status_opcoes else 0,
                        key=f"status_edit_{reserva_id_edicao}"
                    )
                    if status_pagamento_existe:
                        status_pagamento_input = st.text_input(
                            "Status do pagamento",
                            value=status_pagamento_default,
                            key=f"status_pag_edit_{reserva_id_edicao}"
                        )
                with col_basicos[1]:
                    reserva_status_opcoes = list(STATUS_RESERVA.values())
                    reserva_status_escolhido = st.selectbox(
                        "Status da reserva",
                        reserva_status_opcoes,
                        index=reserva_status_opcoes.index(reserva_status_padrao) if reserva_status_padrao in reserva_status_opcoes else 0,
                        key=f"reserva_status_edit_{reserva_id_edicao}"
                    )
                    meia_diaria_input = st.checkbox(
                        "Aplicar meia diária",
                        value=meia_diaria_default,
                        key=f"meia_diaria_edit_{reserva_id_edicao}"
                    )
                    carro_select_placeholder = st.empty()
                    carro_select_placeholder.info("Selecione as datas para listar veículos disponíveis.")

                col_datas = st.columns(3)
                with col_datas[0]:
                    data_inicio_input = st.date_input(
                        "Data de retirada",
                        #value=st.session_state[data_inicio_key],
                        min_value=date.today(),
                        key=data_inicio_key
                    )
                    data_inicio_prev = st.session_state[data_inicio_prev_key]
                    if data_inicio_input != data_inicio_prev:
                        st.session_state[data_inicio_prev_key] = data_inicio_input
                        if st.session_state[data_fim_auto_key]:
                            nova_devolucao = data_inicio_input + timedelta(days=1)
                            st.session_state[data_fim_key] = nova_devolucao
                            st.session_state[data_fim_prev_key] = nova_devolucao
                    else:
                        st.session_state[data_inicio_prev_key] = data_inicio_input
                with col_datas[1]:
                    if st.session_state[data_fim_key] < data_inicio_input:
                        st.session_state[data_fim_key] = data_inicio_input
                        st.session_state[data_fim_prev_key] = data_inicio_input
                    data_fim_input = st.date_input(
                        "Data de devolução",
                        #value=st.session_state[data_fim_key],
                        min_value=data_inicio_input,
                        key=data_fim_key
                    )
                    data_fim_prev = st.session_state[data_fim_prev_key]
                    if data_fim_input != data_fim_prev:
                        st.session_state[data_fim_auto_key] = False
                    st.session_state[data_fim_prev_key] = data_fim_input
                with col_datas[2]:
                    horario_input = st.time_input(
                        "Horário de entrega",
                        value=horario_padrao,
                        key=f"horario_edit_{reserva_id_edicao}"
                    )

                datas_alteradas = (data_inicio_input != data_inicio_reserva or data_fim_input != data_fim_reserva)
                carros_disponiveis_edit = pd.DataFrame()
                carro_escolhido = None

                if data_inicio_input and data_fim_input:
                    carro_select_placeholder.empty()
                    if datas_alteradas:
                        carros_disponiveis_edit = get_available_vehicles(data_inicio_input, data_fim_input)

                        if check_vehicle_availability(carro_atual_id, data_inicio_input, data_fim_input, reserva_id_edicao):
                            carro_atual_info = carros_df[carros_df['id'] == carro_atual_id]
                            if not carro_atual_info.empty:
                                carros_disponiveis_edit = pd.concat([carros_disponiveis_edit, carro_atual_info], ignore_index=True)
                    else:
                        carros_disponiveis_edit = carros_df.copy()

                    if not carros_disponiveis_edit.empty:
                        carros_disponiveis_edit = carros_disponiveis_edit.drop_duplicates(subset='id', keep='first')
                        carros_opcoes = format_vehicle_options(carros_disponiveis_edit)
                        try:
                            idx_carro = next(i for i, op in enumerate(carros_opcoes) if op.startswith(f"{carro_atual_id} -"))
                        except StopIteration:
                            idx_carro = 0

                        carro_escolhido = carro_select_placeholder.selectbox(
                            "Veículo disponível para o período selecionado",
                            carros_opcoes,
                            index=idx_carro,
                            key=carro_select_key
                        )

                        if carro_escolhido:
                            carro_id = int(carro_escolhido.split(" - ")[0])
                            if check_vehicle_availability(carro_id, data_inicio_input, data_fim_input, reserva_id_edicao):
                                st.success("✅ Veículo disponível para o período selecionado")
                            else:
                                st.error("❌ Veículo não disponível para o período selecionado")
                    else:
                        st.session_state[carro_select_key] = ""
                        carro_select_placeholder.warning("Nenhum veículo disponível para o período informado. Ajuste as datas para continuar.")
                else:
                    carro_select_placeholder.info("Selecione as datas para listar veículos disponíveis.")

                carro_escolhido = st.session_state.get(carro_select_key)

                col_km = st.columns(3)
                km_saida_input = col_km[0].number_input(
                    "KM de saída",
                    min_value=0,
                    value=km_saida_default,
                    step=10,
                    key=f"km_saida_edit_{reserva_id_edicao}"
                )
                km_volta_input = col_km[1].number_input(
                    "KM de retorno (se já registrado)",
                    min_value=km_saida_input,
                    value=max(km_volta_default, km_saida_input),
                    step=10,
                    key=f"km_volta_edit_{reserva_id_edicao}"
                )
                km_franquia_input = col_km[2].number_input(
                    "Franquia de KM",
                    min_value=0,
                    value=km_franquia_default,
                    step=10,
                    key=f"km_franquia_edit_{reserva_id_edicao}"
                )

                col_fin1 = st.columns(3)
                adiantamento_input = col_fin1[0].number_input(
                    "Adiantamento (R$)",
                    min_value=0.0,
                    value=adiantamento_default,
                    step=10.0,
                    format="%.2f",
                    key=f"adiantamento_edit_{reserva_id_edicao}"
                )
                pagamento_parcial_input = col_fin1[1].number_input(
                    "Pagamento parcial na entrega (R$)",
                    min_value=0.0,
                    value=pagamento_parcial_default,
                    step=10.0,
                    format="%.2f",
                    key=f"pag_parcial_edit_{reserva_id_edicao}"
                )
                
                col_fin2 = st.columns(3)
                valor_restante_input = col_fin2[0].number_input(
                    "Valor restante (R$)",
                    min_value=0.0,
                    value=valor_restante_default,
                    step=10.0,
                    format="%.2f",
                    key=f"valor_restante_edit_{reserva_id_edicao}",
                    disabled=True
                )
                
                # Inicializar variáveis que serão usadas nas verificações
                valor_diarias_calculado = valor_diarias_default
                total_diarias_calculado = total_diarias_default
                
                # Definir inputs que serão usados na verificação antes de usá-los
                desconto_input = col_fin2[2].number_input(
                    "Desconto concedido (R$)",
                    min_value=0.0,
                    value=desconto_default,
                    step=10.0,
                    format="%.2f",
                    key=f"desconto_edit_{reserva_id_edicao}"
                )
                
                # Verificar se campos que afetam o cálculo foram alterados (antes de usar nos inputs)
                meia_diaria_alterada = (meia_diaria_input != meia_diaria_default)
                desconto_alterado = (desconto_input != desconto_default)
                campos_calculo_alterados = datas_alteradas or meia_diaria_alterada or desconto_alterado
                
                # Recalcular total de diárias se datas, meia diária ou desconto foram alterados e veículo foi selecionado
                if campos_calculo_alterados and carro_escolhido and data_inicio_input and data_fim_input:
                    try:
                        carro_id_selecionado = int(carro_escolhido.split(" - ")[0])
                        carro_info_df = carros_disponiveis_edit[carros_disponiveis_edit['id'] == carro_id_selecionado]
                        if not carro_info_df.empty:
                            diaria_veiculo = float(carro_info_df.iloc[0]['diaria'])
                            dias_periodo = (data_fim_input - data_inicio_input).days 
                            dias_periodo = max(1, dias_periodo)
                            
                            # Calculate valor_diarias (dias * valor diaria) - sem desconto
                            if meia_diaria_input and dias_periodo > 0:
                                valor_diarias_calculado = diaria_veiculo * (dias_periodo - 1) + (diaria_veiculo * 0.5)
                            else:
                                valor_diarias_calculado = diaria_veiculo * dias_periodo
                            
                            # valor_diarias_calculado já contém o valor total sem desconto
                            total_diarias_calculado = valor_diarias_calculado

                            st.info(f"📊 Total de diárias recalculado: R$ {total_diarias_calculado:.2f} ({dias_periodo} dias)")
                    except (ValueError, IndexError, KeyError):
                        pass

                total_diarias_input = col_fin2[1].number_input(
                "Total de diárias (R$)" + (" 📊" if campos_calculo_alterados else ""),
                min_value=0.0,
                value=total_diarias_calculado if campos_calculo_alterados else total_diarias_default,
                format="%.2f",
                key=f"total_diarias_edit_{reserva_id_edicao}",
                disabled=True,
                help="Valor calculado automaticamente: diária × número de dias"
                )

                col_fin3 = st.columns(3)
                valor_multas_input = col_fin3[0].number_input(
                    "Multas (R$)",
                    min_value=0.0,
                    value=valor_multas_default,
                    step=10.0,
                    format="%.2f",
                    key=f"multas_edit_{reserva_id_edicao}"
                )
                valor_danos_input = col_fin3[1].number_input(
                    "Danos (R$)",
                    min_value=0.0,
                    value=valor_danos_default,
                    step=10.0,
                    format="%.2f",
                    key=f"danos_edit_{reserva_id_edicao}"
                )
                valor_outros_input = col_fin3[2].number_input(
                    "Outros custos (R$)",
                    min_value=0.0,
                    value=valor_outros_default,
                    step=10.0,
                    format="%.2f",
                    key=f"outros_edit_{reserva_id_edicao}"
                )

                col_fin4 = st.columns(2)
                custo_lavagem_input = col_fin4[0].number_input(
                    "Custo de lavagem (R$)",
                    min_value=0.0,
                    value=custo_lavagem_default,
                    step=10.0,
                    format="%.2f",
                    key=f"lavagem_edit_{reserva_id_edicao}"
                )
                if observacoes_existe:
                    observacoes_input = col_fin4[1].text_area(
                        "Observações",
                        value=observacoes_default,
                        key=f"observacoes_edit_{reserva_id_edicao}"
                    )

                salvar_alt = st.button(
                    "💾 Salvar alterações",
                    type="primary",
                    width='stretch',
                    key=f"salvar_edit_{reserva_id_edicao}"
                )

                if salvar_alt:
                    if data_fim_input < data_inicio_input:
                        st.error("A data de devolução precisa ser posterior à data de retirada.")
                        st.stop()

                    cliente_id_novo = int(cliente_escolhido.split(" - ")[0])
                    carro_id_novo = int(carro_escolhido.split(" - ")[0])
                    
                    # Verificar disponibilidade do veículo antes de salvar
                    if not check_vehicle_availability(carro_id_novo, data_inicio_input, data_fim_input, reserva_id_edicao):
                        st.error("❌ Não é possível salvar: o veículo não está disponível para o período selecionado.")
                        st.stop()
                    
                    # Calcular valor total automaticamente antes de atualizar
                    dados_carro_update = run_query(
                        "SELECT diaria FROM carros WHERE id = %s", 
                        (carro_id_novo,),
                        fetch=True
                    )
                    if not dados_carro_update.empty:
                        dados_carro_update = dados_carro_update.iloc[0]
                        dias_locacao_edit = (data_fim_input - data_inicio_input).days + 0
                        dias_locacao_edit = max(1, dias_locacao_edit)
                        # Calcular subtotal (diárias + multas + danos + outros + lavagem)
                        subtotal_calculado = total_diarias_calculado + valor_multas_input + valor_danos_input + valor_outros_input + custo_lavagem_input
                        # Aplicar desconto sobre o valor total
                        valor_total_calculado = max(0.0, subtotal_calculado - desconto_input)
                    else:
                        valor_total_calculado = max(0.0, valor_multas_input + valor_danos_input + valor_outros_input + custo_lavagem_input - desconto_input)
                        
                    campos_update = {
                        "cliente_id": cliente_id_novo,
                        "carro_id": carro_id_novo,
                        "data_inicio": data_inicio_input,
                        "data_fim": data_fim_input,
                        "horario_entrega": horario_input,
                        "status": status_escolhido,
                        "reserva_status": reserva_status_escolhido,
                        "km_saida": km_saida_input,
                        "km_volta": km_volta_input,
                        "km_franquia": km_franquia_input,
                        "custo_lavagem": custo_lavagem_input,
                        "adiantamento": adiantamento_input,
                        "pagamento_parcial_entrega": pagamento_parcial_input,
                        "valor_multas": valor_multas_input,
                        "valor_danos": valor_danos_input,
                        "valor_outros": valor_outros_input,
                        "desconto_cliente": desconto_input,
                        "meia_diaria": meia_diaria_input,
                        "total_diarias": total_diarias_calculado,
                        "valor_total": valor_total_calculado,
                        "valor_restante": valor_total_calculado - adiantamento_input - pagamento_parcial_input
                    }
                    if status_pagamento_existe:
                        campos_update["status_pagamento"] = status_pagamento_input
                    if observacoes_existe:
                        campos_update["observacoes"] = observacoes_input

                    set_clause = ", ".join([f"{col}=%s" for col in campos_update.keys()])
                    valores = list(campos_update.values())

                    conn = None
                    try:
                        conn = get_db_connection()
                        conn.autocommit = False
                        cursor = conn.cursor()
                        cursor.execute(
                            f"UPDATE reservas SET {set_clause} WHERE id=%s",
                            (*valores, reserva_id_edicao)
                        )

                        if cursor.rowcount == 0:
                            raise Exception("Nenhuma linha foi atualizada.")

                        status_carro_destino = STATUS_CARRO['DISPONIVEL']
                        if reserva_status_escolhido == STATUS_RESERVA['RESERVADA']:
                            status_carro_destino = STATUS_CARRO['RESERVADO']
                        elif reserva_status_escolhido == STATUS_RESERVA['LOCADA']:
                            status_carro_destino = STATUS_CARRO['LOCADO']

                        cursor.execute(
                            "UPDATE carros SET status=%s WHERE id=%s",
                            (status_carro_destino, carro_id_novo)
                        )

                        if carro_id_novo != carro_atual_id:
                            cursor.execute(
                                "UPDATE carros SET status=%s WHERE id=%s",
                                (STATUS_CARRO['DISPONIVEL'], carro_atual_id)
                            )

                        conn.commit()
                        st.toast("Reserva atualizada com sucesso!", icon="✅")
                        st.success(f"Reserva #{reserva_id_edicao} atualizada.")
                        time.sleep(1)
                        st.rerun()
                    except Exception as e:
                        if conn:
                            conn.rollback()
                        st.error(f"Erro ao salvar alterações: {e}")
                    finally:
                        if conn:
                            conn.close()

                st.markdown("---")
                gerar_pdf = st.button("📄 Gerar contrato atualizado", key=f"btn_contrato_{reserva_id_edicao}")
                if gerar_pdf:
                    if dados_reserva.get('reserva_status') != STATUS_RESERVA['LOCADA']:
                        st.error("O contrato só pode ser gerado após a entrega efetiva do veículo (status 'Locada').")
                    else:
                        # OTIMIZADO: Um único JOIN para buscar cliente e carro
                        dados_contrato_df = run_query_dataframe(
                            """
                            SELECT 
                                cl.*,
                                c.*,
                                r.id as reserva_id
                            FROM reservas r
                            JOIN clientes cl ON r.cliente_id = cl.id
                            JOIN carros c ON r.carro_id = c.id
                            WHERE r.id = %s
                            """,
                            (reserva_id_edicao,)
                        )
                        
                        if not dados_contrato_df.empty:
                            row = dados_contrato_df.iloc[0]
                            
                            # Extrai dados do cliente
                            cliente_dict = {
                                'id': row['id'],
                                'nome': row['nome'],
                                'cpf': row['cpf'],
                                'rg': row['rg'],
                                'cnh': row['cnh'],
                                'validade_cnh': row['validade_cnh'],
                                'telefone': row['telefone'],
                                'endereco': row['endereco'],
                                'status': row['status']
                            }
                            
                            # Extrai dados do carro
                            carro_dict = {
                                'id': row['id'],
                                'marca': row['marca'],
                                'modelo': row['modelo'],
                                'placa': row['placa'],
                                'cor': row['cor'],
                                'diaria': row['diaria'],
                                'preco_km': row['preco_km'],
                                'km_atual': row['km_atual'],
                                'status': row['status'],
                                'numero_chassi': row['numero_chassi'],
                                'numero_renavam': row['numero_renavam'],
                                'ano_veiculo': row['ano_veiculo'],
                                'km_troca_oleo': row['km_troca_oleo']
                            }
                            pdf_bytes_contrato = gerar_contrato_pdf(
                                cliente_dict,
                                carro_dict,
                                data_inicio_reserva,
                                data_fim_reserva,
                                horario_padrao
                            )
                            st.session_state.pdf_para_download = pdf_bytes_contrato
                            st.session_state.pdf_file_name = f"contrato_reserva_{reserva_id_edicao}.pdf"
                            st.success("Contrato gerado! Utilize o botão abaixo para baixar.")
                        else:
                            st.error("Não foi possível obter dados completos para gerar o contrato.")

        if st.session_state.get('pdf_para_download'):
            st.markdown("---")
            st.download_button(
                label="📥 Baixar Contrato em PDF",
                data=st.session_state.pdf_para_download,
                file_name=st.session_state.pdf_file_name,
                mime="application/pdf",
                key="download_contrato_edicao_reserva"
            )
            if st.button("✅ Concluído - Limpar", key="limpar_download_contrato"):
                st.session_state.pdf_para_download = None
                st.session_state.pdf_file_name = None
                st.rerun()

elif menu == "Entrega do veículo":
    render_section_header(
        title="Entrega Simplificada",
        subtitle="Confirme dados e gerencie reservas antes de liberar o carro.",
        icon="🔑",
        trail=["Reservas", "Entrega"]
    )

    def carregar_dados(id_reserva_sel, carro_id, cliente_id):
        try:
            # OTIMIZADO: Um único JOIN em vez de 3 queries separadas
            query_completa = run_query(
                """
                SELECT 
                    r.*,
                    c.id as carro_id_full, c.marca, c.modelo, c.placa, c.cor, c.diaria, 
                    c.preco_km, c.km_atual, c.status, c.numero_chassi, c.numero_renavam, 
                    c.ano_veiculo, c.km_troca_oleo,
                    cl.id as cliente_id_full, cl.nome, cl.cpf, cl.rg, cl.cnh, cl.validade_cnh, 
                    cl.telefone, cl.endereco, cl.status as status_cliente
                FROM reservas r
                JOIN carros c ON r.carro_id = c.id
                JOIN clientes cl ON r.cliente_id = cl.id
                WHERE r.id = %s
                """,
                (id_reserva_sel,),
                fetch=True
            )
            
            if query_completa is not None and not query_completa.empty:
                row = query_completa.iloc[0]
                
                # Extrai dados do carro
                dados_carro = {
                    'id': row['carro_id_full'],
                    'marca': row['marca'],
                    'modelo': row['modelo'],
                    'placa': row['placa'],
                    'cor': row['cor'],
                    'diaria': row['diaria'],
                    'preco_km': row['preco_km'],
                    'km_atual': row['km_atual'],
                    'status': row['status'],
                    'numero_chassi': row['numero_chassi'],
                    'numero_renavam': row['numero_renavam'],
                    'ano_veiculo': row['ano_veiculo'],
                    'km_troca_oleo': row['km_troca_oleo']
                }
                
                # Extrai dados do cliente
                dados_cliente = {
                    'id': row['cliente_id_full'],
                    'nome': row['nome'],
                    'cpf': row['cpf'],
                    'rg': row['rg'],
                    'cnh': row['cnh'],
                    'validade_cnh': row['validade_cnh'],
                    'telefone': row['telefone'],
                    'endereco': row['endereco'],
                    'status': row['status_cliente']
                }
                
                # Extrai dados da reserva
                dados_reserva = {
                    'id': row['id'],
                    'carro_id': row['carro_id'],
                    'cliente_id': row['cliente_id'],
                    'data_inicio': row['data_inicio'],
                    'data_fim': row['data_fim'],
                    'status': row['status'],
                    'reserva_status': row['reserva_status'],
                    'km_saida': row['km_saida'],
                    'km_franquia': row['km_franquia'],
                    'adiantamento': row['adiantamento'],
                    'valor_multas': row['valor_multas'],
                    'valor_danos': row['valor_danos'],
                    'valor_outros': row['valor_outros'],
                    'desconto_cliente': row['desconto_cliente'],
                    'meia_diaria': row['meia_diaria'],
                    'total_diarias': row['total_diarias']
                }
                
                return dados_carro, dados_cliente, dados_reserva
        except Exception as e:
            st.error(f"Erro ao carregar dados: {e}")
        return None, None, None

    def finalizar_entrega(carro_id, id_reserva_sel, km_confirma, data_saida, horario_entrega,
                          dados_cliente, dados_carro, dados_reserva, adiantamento=0.0, valor_restante=0.0):
        """
        Atualiza banco e gera contrato ao liberar o veículo para o cliente.
        """
        conn = None
        try:
            conn = get_db_connection()
            conn.autocommit = False
            cursor = conn.cursor()

            st.toast("Iniciando processamento da entrega...", icon="⏳")

            try:
                st.toast("Atualizando status do veículo...", icon="🔄")
                cursor.execute(
                    "UPDATE carros SET status=%s, km_atual=%s WHERE id=%s RETURNING id",
                    (STATUS_CARRO['LOCADO'], km_confirma, carro_id)
                )
                if cursor.rowcount == 0:
                    raise Exception(f"Carro com ID {carro_id} não encontrado")

                st.toast("Atualizando dados da reserva...", icon="📝")
                dias_locacao = (pd.to_datetime(dados_reserva['data_fim']).date() - data_saida).days + 0
                dias_locacao = max(1, dias_locacao)
                valor_total = float(dados_carro['diaria']) * dias_locacao

                cursor.execute("""
                    UPDATE reservas
                    SET km_saida=%s,
                        data_inicio=%s,
                        reserva_status=%s,
                        horario_entrega=%s,
                        adiantamento=%s,
                        valor_total=%s,
                        valor_restante=%s
                    WHERE id=%s
                    RETURNING id
                """, (
                    km_confirma,
                    data_saida,
                    STATUS_RESERVA['LOCADA'],
                    horario_entrega,
                    float(adiantamento),
                    valor_total,
                    float(valor_restante),
                    id_reserva_sel
                ))

                if cursor.rowcount == 0:
                    raise Exception(f"Falha ao atualizar a reserva ID {id_reserva_sel}")

                st.toast("Gerando contrato...", icon="📄")
                try:
                    pdf_bytes = gerar_contrato_pdf(
                        dados_cliente,
                        dados_carro,
                        data_saida,
                        pd.to_datetime(dados_reserva['data_fim']).date(),
                        horario_entrega
                    )
                except Exception as e:
                    raise Exception(f"Erro ao gerar contrato: {e}")

                conn.commit()
                st.toast("Entrega processada com sucesso!", icon="✅")
                return True, "Entrega processada com sucesso!", pdf_bytes

            except Exception as e:
                if conn:
                    conn.rollback()
                return False, f"Erro durante o processamento: {str(e)}", None

        except Exception as e:
            return False, f"Erro inesperado: {str(e)}", None
        finally:
            if conn:
                try:
                    conn.close()
                except Exception:
                    pass

    def finalizar_entrega_simples(**kwargs):
        return finalizar_entrega(
            carro_id=kwargs['carro_id'],
            id_reserva_sel=kwargs['id_reserva_sel'],
            km_confirma=kwargs['km_confirma'],
            data_saida=kwargs['data_saida'],
            horario_entrega=kwargs['horario_entrega'],
            dados_cliente=kwargs['dados_cliente'],
            dados_carro=kwargs['dados_carro'],
            dados_reserva=kwargs['dados_reserva'],
            adiantamento=kwargs['adiantamento'],
            valor_restante=kwargs['valor_restante']
        )

    if 'pdf_para_download' not in st.session_state:
        st.session_state.pdf_para_download = None
        st.session_state.pdf_file_name = None

    # Otimizado: Uma única query com todos os dados necessários
    reservas_entrega_df = get_reservas_entrega()

    if reservas_entrega_df.empty:
        st.info("Nenhuma reserva aguardando entrega.")
    else:
        opcoes = ["Selecione a reserva..."] + reservas_entrega_df.apply(
            lambda r: f"ID {r['id']} • {r['cliente_nome']} ({r['modelo']} - {r['placa']} - Início {r['data_inicio']} → {r['data_fim']})",
            axis=1
        ).tolist()
        escolha = st.selectbox("Reserva pronta para entrega", opcoes, key="entrega_simples_sel")

        if escolha != "Selecione a reserva...":
            reserva_id = int(escolha.split(" ")[1])
            reserva_row = reservas_entrega_df[reservas_entrega_df['id'] == reserva_id].iloc[0]
            dados_carro, dados_cliente, dados_reserva = carregar_dados(
                reserva_id,
                int(reserva_row['carro_id']),
                int(reserva_row['cliente_id'])
            )

            if None not in (dados_carro, dados_cliente, dados_reserva):
                if validar_cnh_simplificada(dados_cliente):
                    km_saida = int(dados_reserva.get('km_saida') or dados_carro.get('km_atual') or 0)
                    valor_previsto = float(dados_reserva.get('total_diarias') or 0.0)
                    if valor_previsto == 0:
                        dias_previstos = max(1, (pd.to_datetime(dados_reserva['data_fim']) - pd.to_datetime(dados_reserva['data_inicio'])).days + 1)
                        valor_previsto = dias_previstos * float(dados_carro.get('diaria') or 0.0)
                    adiantamento_reserva = float(dados_reserva.get('adiantamento') or 0.0)
                    saldo = max(0.0, valor_previsto - adiantamento_reserva)

                    col_summary = st.columns(4)
                    col_summary[0].metric("Cliente", dados_cliente.get('nome', '—'))
                    col_summary[1].metric("Veículo", f"{dados_carro.get('marca', '')} {dados_carro.get('modelo', '')}")
                    col_summary[2].metric(
                        "Período",
                        f"{pd.to_datetime(dados_reserva['data_inicio']).date():%d/%m} → {pd.to_datetime(dados_reserva['data_fim']).date():%d/%m}"
                    )
                    col_summary[3].metric("Saldo Devedor", f"R$ {saldo:.2f}")

                    with st.form("form_entrega_simplificada"):
                        km_confirma = st.number_input(
                            "KM conferido na saída",
                            min_value=km_saida,
                            value=km_saida,
                            key="entrega_simples_km"
                        )
                        data_saida_real = st.date_input(
                            "Data real da saída",
                            #value=datetime.now().date(),
                            value=pd.to_datetime(dados_reserva['data_inicio']).date(),
                            key="entrega_simples_data"
                        )
                        horario_entrega = st.time_input(
                            "Horário da entrega",
                            value=datetime.strptime("09:00", "%H:%M").time(),
                            key="entrega_simples_horario"
                        )
                        valor_pago_agora = st.number_input(
                            "Valor recebido agora (R$)",
                            min_value=0.0,
                            max_value=max(saldo, 0.0),
                            value=saldo,
                            step=10.0,
                            format="%.2f",
                            key="entrega_simples_valor"
                        )
                        submit = st.form_submit_button("Confirmar entrega", type="primary", width='stretch')

                    if submit:
                        total_pago = adiantamento_reserva + valor_pago_agora
                        valor_restante = max(0.0, valor_previsto - total_pago)

                        dados_carro['km_atual'] = km_confirma  # Atualiza a quilometragem do carro
                        sucesso, mensagem, pdf_bytes = finalizar_entrega_simples(
                            carro_id=int(reserva_row['carro_id']),
                            id_reserva_sel=reserva_id,
                            km_confirma=km_confirma,
                            data_saida=data_saida_real,
                            horario_entrega=horario_entrega,
                            dados_cliente=dados_cliente,
                            dados_carro=dados_carro,
                            dados_reserva=dados_reserva,
                            adiantamento=total_pago,
                            valor_restante=valor_restante
                        )

                        if sucesso and pdf_bytes:
                            st.session_state.pdf_para_download = pdf_bytes
                            data_atual = date.today()
                            nome_formatado = dados_cliente['nome'].replace(' ', '_').lower()
                            modelo_formatado = dados_carro['modelo'].replace(' ', '_').lower()
                            st.session_state.pdf_file_name = f"contrato_{nome_formatado}_{modelo_formatado}_{data_atual.day}_{data_atual.month}.pdf"
                            st.success("Entrega confirmada! Baixe o contrato abaixo.")
                        else:
                            st.error(f"Erro ao finalizar entrega: {mensagem}")

                    if st.session_state.pdf_para_download:
                        st.download_button(
                            "📄 Baixar contrato",
                            data=st.session_state.pdf_para_download,
                            file_name=st.session_state.pdf_file_name,
                            mime="application/pdf",
                            key="download_contrato_simples"
                        )

    
# 6. DEVOLUÇÃO
elif menu == "Devolução":
    render_section_header(
        title="Devolução de Veículo",
        subtitle="Finalize a locação registrando KM, custos extras e recibos.",
        icon="🔄",
        trail=["Operação", "Devolução"]
    )

    # Inicializa o Session State para o PDF, se necessário
    if 'pdf_para_download' not in st.session_state:
        st.session_state.pdf_para_download = None
        st.session_state.pdf_file_name = None

    # MUDANÇA NO FILTRO: Busca reservas com reserva_status='Locada' (Carro em uso pelo cliente)
    query_dev = """
    SELECT 
        r.id, cl.nome, cl.cpf, cl.telefone, cl.endereco, c.modelo, c.placa, r.km_saida, c.preco_km, c.diaria, 
        r.data_inicio, r.carro_id, r.cliente_id, r.km_franquia, r.adiantamento, 
        r.valor_multas, r.valor_danos, r.valor_outros, r.desconto_cliente, r.meia_diaria,
        r.total_diarias
    FROM reservas r 
    JOIN carros c ON r.carro_id = c.id 
    JOIN clientes cl ON r.cliente_id = cl.id
    WHERE r.status='Ativa' AND r.reserva_status='Locada'
    """

    ativas = run_query(query_dev, fetch=True)

    if isinstance(ativas, str):
        st.error(f"Erro no banco de dados: {ativas}")
    elif not ativas.empty:
        opcoes = ativas.apply(lambda x: f"{x['id']} - {x['nome']} ({x['modelo']} - {x['placa']})", axis=1)
        opcoes_com_placeholder = ["Selecione a locação pendente..."] + opcoes.tolist()

        sel = st.selectbox("Selecione a Locação Pendente", opcoes_com_placeholder)

        if sel != "Selecione a locação pendente...":

            try:
                id_reserva_sel = int(sel.split(" - ")[0])
                reserva = ativas[ativas['id'] == id_reserva_sel].iloc[0]
            except:
                st.warning("Erro ao processar ID da reserva. Selecione novamente.")
                reserva = None

            if reserva is not None:

                try:
                    km_saida_safe = int(reserva['km_saida'])
                except:
                    km_saida_safe = 0
                    st.warning("Aviso: KM de saída com formato inválido no banco. Usando 0.")

                # OTIMIZADO: Dados do cliente já estão disponíveis na query anterior (ativas)
                # Não é necessária query adicional - usar dados já carregados
                dados_cliente = {
                    'id': reserva['cliente_id'],
                    'nome': reserva['nome'],
                    'cpf': reserva.get('cpf', ''),
                    'telefone': reserva.get('telefone', ''),
                    'endereco': reserva.get('endereco', '')
                }

                # Cabeçalho com informações principais
                st.markdown("---")
                st.subheader("📋 Dados da Devolução")
                
                # Seção 1: Informações básicas em colunas
                col_info1, col_info2 = st.columns(2)
                
                with col_info1:
                    st.markdown("### Informações do Veículo")
                    st.markdown(f"**Modelo:** {reserva['modelo']}")
                    st.markdown(f"**Placa:** {reserva['placa']}")
                    st.markdown(f"**Data de Retirada:** {pd.to_datetime(reserva['data_inicio']).strftime('%d/%m/%Y')}")
                    st.markdown(f"**KM de Saída:** {km_saida_safe} km")
                    
                    # Campo para KM de devolução
                    km_volta = st.number_input(
                        "🔢 KM de Devolução",
                        min_value=km_saida_safe,
                        value=km_saida_safe,
                        help="A quilometragem não pode ser menor que a da saída.",
                        key="km_devolucao"
                    )
                
                with col_info2:
                    st.markdown("### Informações do Cliente")
                    st.markdown(f"**Nome:** {reserva['nome']}")
                    st.markdown(f"**CPF:** {dados_cliente.get('cpf', 'Não informado')}")
                    st.markdown(f"**Telefone:** {dados_cliente.get('telefone', 'Não informado')}")
                    st.markdown(f"**Data de Devolução:** {date.today().strftime('%d/%m/%Y')}")
                
                st.markdown("---")
                st.subheader("💸 Cálculos da Locação")
                
                # Cálculos
                data_saida_real = pd.to_datetime(reserva['data_inicio']).date()
                data_devolucao = date.today()
                if data_devolucao < data_saida_real:
                    st.warning(
                        "Data de devolução ajustada para a data de saída registrada, "
                        "pois não é permitido finalizar antes da retirada."
                    )
                    data_devolucao = data_saida_real
                dias = (data_devolucao - data_saida_real).days
                dias_cobranca = max(dias, 1)  # Mínimo 1 dia

                # Cálculo de KM rodados
                km_rodados_totais = km_volta - km_saida_safe
                km_franquia_reserva = reserva['km_franquia'] if reserva['km_franquia'] is not None else 0
                
                if km_rodados_totais > km_franquia_reserva:
                    km_franquia_reserva = 0
                    
                km_a_cobrar = max(0, km_rodados_totais - km_franquia_reserva)
                custo_km = km_a_cobrar * reserva['preco_km']
                
                # Usar total_diarias armazenado da reserva em vez de recalcular
                valor_diarias_stored = Decimal(str(reserva.get('total_diarias', 0.0) or 0.0))
                total_diarias_stored = Decimal(str(reserva.get('total_diarias', 0.0) or 0.0))
                
                # Usar total_diarias que já inclui o desconto aplicado na criação da reserva
                custo_diarias_com_desconto = total_diarias_stored
                
                # Seção 2: Custos adicionais
                with st.expander("➕ Adicionar Custos Extras", expanded=False):
                    st.markdown("### Custos Adicionais")
                    
                    col_extra1, col_extra2 = st.columns(2)
                    
                    with col_extra1:
                        st.markdown("#### Serviços")
                        cobrar_lavagem = st.checkbox("Adicionar Lavagem", value=False, key="cobrar_lavagem")
                        valor_lavagem = 0.0
                        if cobrar_lavagem:
                            valor_lavagem = st.number_input(
                                "Valor da Lavagem (R$)", 
                                value=50.0, 
                                min_value=0.0, 
                                step=5.0, 
                                key="valor_lavagem"
                            )
                    
                    with col_extra2:
                        st.markdown("#### Outros Custos")
                        valor_multas = st.number_input(
                            "Multas (R$)", 
                            min_value=0.0, 
                            value=0.0, 
                            step=10.0, 
                            format="%.2f", 
                            key="valor_multas"
                        )
                        valor_danos = st.number_input(
                            "Danos ao Veículo (R$)", 
                            min_value=0.0, 
                            value=0.0, 
                            step=10.0, 
                            format="%.2f", 
                            key="valor_danos"
                        )
                        valor_outros = st.number_input(
                            "Outros Custos (R$)", 
                            min_value=0.0, 
                            value=0.0, 
                            step=10.0, 
                            format="%.2f", 
                            key="valor_outros"
                        )
                # Cálculos finais
                subtotal_sem_adiantamento = (
                    custo_diarias_com_desconto + 
                    Decimal(str(custo_km)) + 
                    Decimal(str(valor_lavagem)) + 
                    Decimal(str(valor_multas)) + 
                    Decimal(str(valor_danos)) + 
                    Decimal(str(valor_outros))
                )
                total_final = subtotal_sem_adiantamento - Decimal(str(reserva.get('adiantamento', 0.0)))
                
                # Inicializa as variáveis de pagamento
                valor_restante = max(Decimal('0'), total_final)  # Inicializa o valor restante
                valor_pago = Decimal('0.0')  # Inicializa o valor pago
                
                # Define o rótulo do total
                label_total = "Total a Pagar (R$)" if total_final >= 0 else "Valor a Devolver ao Cliente (R$)"
                valor_display = formatar_moeda(abs(total_final))
                
                # Seção 3: Resumo dos cálculos
                st.markdown("---")
                st.subheader("📊 Resumo Financeiro")
                
                # Tabela de resumo
                resumo_data = {
                    'Descrição': [
                        f"Diárias ({dias_cobranca} dias) - Valor base: {formatar_moeda(valor_diarias_stored)}",
                        f"KM Rodados ({km_rodados_totais} km - {km_a_cobrar} km cobráveis)",
                        "Lavagem" if valor_lavagem > 0 else None,
                        "Multas" if valor_multas > 0 else None,
                        "Danos ao Veículo" if valor_danos > 0 else None,
                        "Outros Custos" if valor_outros > 0 else None,
                        "**Subtotal**",
                        "(-) Adiantamento Pago",
                        f"**{label_total}**"
                    ],
                    'Valor (R$)': [
                        formatar_moeda(custo_diarias_com_desconto),
                        formatar_moeda(custo_km),
                        formatar_moeda(valor_lavagem) if valor_lavagem > 0 else None,
                        formatar_moeda(valor_multas) if valor_multas > 0 else None,
                        formatar_moeda(valor_danos) if valor_danos > 0 else None,
                        formatar_moeda(valor_outros) if valor_outros > 0 else None,
                        f"**{formatar_moeda(subtotal_sem_adiantamento)}**",
                        f"-{formatar_moeda(reserva['adiantamento'])}",
                        f"**{valor_display}**"
                    ]
                }
                
                # Remove linhas vazias
                resumo_data = {k: [v for v in vs if v is not None] for k, vs in resumo_data.items()}
                
                # Exibe a tabela de resumo
                st.table(pd.DataFrame(resumo_data))
                
                # Seção 4: Pagamento
                st.markdown("---")
                st.subheader("💳 Pagamento")
                
                if total_final > 0:
                    # Se há valor a pagar
                    st.info(f"Valor a ser pago: **{formatar_moeda(total_final)}**")
                    
                    # Campo para valor pago
                    valor_pago = st.number_input(
                        "Valor Recebido (R$)",
                        min_value=0.0,
                        max_value=float(total_final * 2),  # Convert to float for Streamlit compatibility
                        value=float(total_final) if total_final > 0 else 0.0,
                        step=1.0,
                        format="%.2f",
                        key="valor_pago_devolucao"
                    )
                    
                    # Calcula o troco, se necessário
                    from decimal import Decimal
                    troco = max(Decimal('0'), Decimal(str(valor_pago)) - total_final) if total_final > 0 else Decimal('0')
                    if troco > 0:
                        st.success(f"💰 Troco: {formatar_moeda(troco)}")
                    
                    # Atualiza o valor restante após o pagamento
                    valor_restante = max(Decimal('0'), total_final - Decimal(str(valor_pago)))
                    
                    if valor_restante > 0:
                        st.warning(f"⚠️ Valor pendente: {formatar_moeda(valor_restante)}")
                else:
                    # Se não há valor a pagar (ou há valor a devolver)
                    valor_pago = 0.0
                    valor_restante = 0.0
                    
                    if total_final < 0:
                        st.success(f"✅ Valor a ser devolvido ao cliente: {formatar_moeda(abs(total_final))}")
                    else:
                        st.success("✅ O valor do adiantamento cobre todos os custos. Não há valor adicional a pagar.")
                
                # Botão de finalização
                if st.button("✅ Finalizar Devolução e Liberar Veículo", type="primary", key="btn_finalizar_devolucao"):
                    conn = None
                    try:
                        # Inicia a transação
                        conn = get_db_connection()
                        conn.autocommit = False  # Desativa o autocommit para usar transações
                        cursor = conn.cursor()
                        
                        st.toast("Iniciando processamento da devolução...", icon="⏳")
                        
                        try:
                            # 1. Busca dados atuais da reserva e bloqueia o registro
                            cursor.execute("""
                                SELECT * FROM reservas WHERE id = %s FOR UPDATE
                            """, (id_reserva_sel,))
                            
                            reserva_atual = cursor.fetchone()
                            if not reserva_atual:
                                raise Exception(f"Reserva com ID {id_reserva_sel} não encontrada")
                            
                            # 2. Calcula valores
                            valor_restante_final = max(0, float(valor_restante)) if total_final > 0 else 0.0
                            valor_pago_efetivo = min(valor_pago, total_final) if total_final > 0 else 0.0
                            
                            # 3. Atualiza a reserva
                            st.toast("Atualizando dados da reserva...", icon="📝")
                            
                            # Converte valores numpy para tipos nativos do Python
                            def to_python_value(value):
                                if hasattr(value, 'item'):  # Para numpy types
                                    return value.item()
                                return value
                                
                            cursor.execute("""
                                UPDATE reservas
                                SET status = %s, 
                                    reserva_status = %s, 
                                    km_volta = %s, 
                                    custo_lavagem = %s, 
                                    valor_total = %s,
                                    valor_multas = %s, 
                                    valor_danos = %s, 
                                    valor_outros = %s, 
                                    total_diarias = %s,
                                    valor_restante = %s,
                                    data_fim = %s
                                WHERE id = %s
                                RETURNING id
                            """, (
                                'Finalizada',
                                'Finalizada',
                                to_python_value(km_volta), 
                                to_python_value(valor_lavagem), 
                                to_python_value(subtotal_sem_adiantamento),
                                to_python_value(valor_multas),
                                to_python_value(valor_danos),
                                to_python_value(valor_outros),
                                to_python_value(custo_diarias_com_desconto),
                                to_python_value(valor_restante_final),
                                data_devolucao,
                                int(id_reserva_sel)  # Garante que o ID seja um inteiro
                            ))
                            
                            if cursor.rowcount == 0:
                                raise Exception("Falha ao atualizar a reserva")
                            
                            # 4. Atualiza o status do carro para disponível
                            st.toast("Atualizando status do veículo...", icon="🔄")
                            cursor.execute("""
                                UPDATE carros 
                                SET status = %s,
                                    km_atual = %s
                                WHERE id = %s
                                RETURNING id
                            """, (
                                STATUS_CARRO['DISPONIVEL'],
                                to_python_value(km_volta),
                                int(reserva['carro_id'])  # Garante que o ID seja um inteiro
                            ))
                            
                            if cursor.rowcount == 0:
                                raise Exception("Falha ao atualizar o status do veículo")
                            
                            # 5. Busca dados completos do carro para o recibo
                            cursor.execute("SELECT * FROM carros WHERE id = %s", (int(reserva['carro_id']),))
                            carro_recibo = cursor.fetchone()
                            
                            if carro_recibo:
                                # Converte o resultado para dicionário
                                colunas = [desc[0] for desc in cursor.description]
                                dados_carro_recibo = dict(zip(colunas, carro_recibo))
                            else:
                                # Se não encontrar o carro, usa os dados da reserva
                                dados_carro_recibo = {
                                    'modelo': reserva['modelo'],
                                    'placa': reserva['placa'],
                                    'cor': 'Não informada',
                                    'ano': 'Não informado',
                                    'km_atual': km_volta,
                                    'chassi': 'Não informado',
                                    'renavam': 'Não informado'
                                }
                            
                            # 6. Prepara os dados para o recibo
                            recibo_dados = {
                                'data_inicio': data_saida_real,
                                'data_fim': data_devolucao,
                                'km_saida': km_saida_safe,
                                'km_volta': km_volta,
                                'km_rodados': km_rodados_totais,
                                'km_franquia': km_franquia_reserva,
                                'km_excedente': max(0, km_rodados_totais - km_franquia_reserva),
                                'dias_cobranca': dias_cobranca,
                                'valor_pago': valor_pago_efetivo,
                                'valor_restante': valor_restante_final,
                                'custo_diarias': custo_diarias_com_desconto,
                                'custo_km': custo_km,
                                'valor_lavagem': valor_lavagem,
                                'valor_multas': valor_multas,
                                'valor_danos': valor_danos,
                                'valor_outros': valor_outros,
                                'adiantamento': reserva['adiantamento'] if reserva['adiantamento'] is not None else 0.0,
                                'total_geral': total_final,
                                'total_final': total_final  # Adicionado para compatibilidade
                            }
                            
                            # 7. Gera o recibo
                            st.toast("Gerando recibo...", icon="📄")
                            recibo_pdf_bytes = gerar_recibo_pdf(dados_cliente, dados_carro_recibo, recibo_dados)
                            
                            # 8. Confirma a transação
                            conn.commit()
                            
                            # Salva o PDF no session state para download
                            st.session_state.pdf_para_download = recibo_pdf_bytes
                            st.session_state.pdf_file_name = f"recibo_devolucao_{reserva['placa']}_{date.today().strftime('%Y%m%d')}.pdf"
                            
                            st.toast("✅ Devolução processada com sucesso!", icon="✅")
                            st.success(f"Devolução do veículo {reserva['placa']} registrada com sucesso!")
                            st.balloons()
                            
                            # Força atualização da página para limpar o formulário
                            # O st.rerun() foi movido para depois do download do PDF
                            
                        except Exception as e:
                            # Em caso de erro, faz rollback
                            if conn:
                                conn.rollback()
                            st.toast("❌ Erro ao processar devolução", icon="❌")
                            st.error(f"Erro durante o processamento: {str(e)}")
                            st.exception(e)  # Mostra o traceback completo para depuração
                            
                    except Exception as e:
                        st.toast("❌ Erro inesperado", icon="❌")
                        st.error(f"Erro inesperado ao processar devolução: {str(e)}")
                        st.exception(e)  # Mostra o traceback completo para depuração
                        
                    finally:
                        # Garante que a conexão será fechada
                        if conn:
                            try:
                                conn.close()
                            except Exception as e:
                                st.error(f"Erro ao fechar conexão com o banco de dados: {str(e)}")
                    carro_id_recibo = reserva['carro_id']
                    if carro_id_recibo is None or pd.isna(carro_id_recibo):
                        st.error("Erro: ID do carro não encontrado na reserva.")
                        st.stop()
                    dados_carro_recibo = run_query("SELECT * FROM carros WHERE id=%s", (int(carro_id_recibo),), fetch=True).iloc[0].to_dict()
                    dados_carro_recibo['chassi'] = dados_carro_recibo.get('numero_chassi')
                    dados_carro_recibo['renavam'] = dados_carro_recibo.get('numero_renavam')

                    # Geração do Recibo em PDF
                    # Calcula o valor pago e o valor restante
                    valor_pago_efetivo = min(valor_pago, valor_restante) if valor_restante > 0 else 0.0
                    valor_restante_final = max(Decimal('0'), valor_restante - Decimal(str(valor_pago))) if valor_restante > 0 else Decimal('0.0')
                    
                    recibo_pdf_bytes = gerar_recibo_pdf(
                        dados_cliente,
                        dados_carro_recibo,
                        {
                            'data_inicio': data_saida_real,
                            'data_fim': data_devolucao,
                            'km_saida': km_saida_safe,
                            'km_volta': km_volta,
                            'km_franquia': km_franquia_reserva,
                            'dias_cobranca': dias_cobranca,
                            'valor_pago': valor_pago_efetivo,
                            'valor_restante': valor_restante_final,
                            'custo_diarias': custo_diarias_com_desconto,
                            'custo_km': custo_km,
                            'valor_lavagem': valor_lavagem,
                            'valor_multas': valor_multas,
                            'valor_danos': valor_danos,
                            'valor_outros': valor_outros,
                            'adiantamento': reserva['adiantamento'] if reserva['adiantamento'] is not None else 0.0,
                            'total_final': total_final
                        }
                    )

                    # Salvar o recibo em Session State para download
                    st.session_state.pdf_para_download = recibo_pdf_bytes
                    st.session_state.pdf_file_name = f"recibo_{dados_cliente['nome']}_{reserva['placa']}_{date.today().strftime('%Y%m%d')}.pdf"

                    st.balloons()
                    #st.rerun()

            if st.session_state.pdf_para_download:
                # Função para limpar o estado após o download
                def limpar_estado():
                    st.session_state.pdf_para_download = None
                    st.session_state.pdf_file_name = None
                
                # Botão de download com callback para limpar após o download
                st.download_button(
                    label="📥 Baixar Comprovante de Devolução",
                    data=st.session_state.pdf_para_download,
                    file_name=st.session_state.pdf_file_name,
                    mime="application/pdf",
                    key="download_recibo_pos_devolucao",
                    on_click=limpar_estado
                )
                

    else:
        st.info("Nenhum veículo em locação para ser devolvido.")

# 7. HISTÓRICO
elif menu == "Histórico":
    st.title("📋 Histórico de Locações")
    
    # Filtros de data
    col1, col2 = st.columns(2)
    data_inicio = col1.date_input("Data Inicial", value=date.today() - timedelta(days=30))
    data_fim = col2.date_input("Data Final", value=date.today())
    
    # Buscar locações finalizadas no período com informações de multas
    query = """
    SELECT 
        r.id,
        c.nome as cliente,
        v.modelo,
        v.placa,
        v.modelo as veiculo,
        r.data_inicio,
        r.data_fim,
        r.valor_total,
        r.km_volta,
        r.km_saida,
        r.status,
        r.valor_restante,
        COALESCE(COUNT(m.id), 0) as quantidade_multas,
        COALESCE(SUM(CASE WHEN m.status != 'Isentada' THEN m.valor ELSE 0 END), 0) as valor_total_multas,
        COALESCE(COUNT(CASE WHEN m.status = 'Paga' THEN 1 END), 0) as multas_pagas
    FROM reservas r
    JOIN clientes c ON r.cliente_id = c.id
    JOIN carros v ON r.carro_id = v.id
    LEFT JOIN multas m ON r.id = m.reserva_id
    WHERE r.status IN ('Finalizada', 'Com Multa Pendente')
    AND r.data_fim BETWEEN %s AND %s + INTERVAL '1 day'
    GROUP BY r.id, c.nome, v.modelo, v.placa, v.modelo, r.data_inicio, r.data_fim, r.valor_total, r.km_volta, r.km_saida, r.status, r.valor_restante
    ORDER BY r.data_fim DESC
    """
    
    df_locacoes = run_query_dataframe(query, params=(data_inicio, data_fim))
    
    if not df_locacoes.empty:
        # Formatar datas e valores para exibição
        df_display = df_locacoes.copy()
        df_display['data_inicio'] = pd.to_datetime(df_display['data_inicio'], format='%Y-%m-%d %H:%M:%S').dt.strftime('%d/%m/%Y %H:%M')
        df_display['data_fim'] = pd.to_datetime(df_display['data_fim'], format='%Y-%m-%d %H:%M:%S').dt.strftime('%d/%m/%Y %H:%M')
        df_display['valor_total'] = df_display['valor_total'].apply(formatar_moeda)
        df_display['valor_total_multas'] = df_display['valor_total_multas'].apply(formatar_moeda)
        df_display['km_rodados'] = df_display['km_volta'] - df_display['km_saida']
        df_display['multas_info'] = df_display.apply(
            lambda row: f"{row['quantidade_multas'] - row.get('multas_pagas', 0)} pendente(s) - {row.get('multas_pagas', 0)} paga(s)" 
                if row['quantidade_multas'] > 0 
                else "Sem multas", 
            axis=1
        )
        df_display['valor_restante'] = df_display['valor_restante'].apply(formatar_moeda)
        
        # Exibir métricas resumidas
        total_locacoes = len(df_locacoes)
        faturamento_total = df_locacoes['valor_total'].sum()
        km_total = (df_locacoes['km_volta'] - df_locacoes['km_saida']).sum()
        
        col1, col2, col3 = st.columns(3)
        col1.metric("Total de Locações", total_locacoes)
        col2.metric("Faturamento Total", formatar_moeda(faturamento_total))
        col3.metric("Quilometragem Total", f"{km_total} km")
        
        # Filtros adicionais
        st.subheader("Filtros")
        col_f1, col_f2, col_f3 = st.columns(3)
        
        # Filtro por tipo de veículo
        tipos_veiculo = ['Todos'] + sorted(df_locacoes['veiculo'].unique().tolist())
        tipo_selecionado = col_f1.selectbox("Tipo de Veículo", tipos_veiculo)
        
        # Aplicar filtros
        if tipo_selecionado != 'Todos':
            df_display = df_display[df_display['veiculo'] == tipo_selecionado]
        
        # Tabela de locações
        st.subheader("Locações Finalizadas")
        st.dataframe(
            df_display[[
                'id', 'cliente', 'modelo', 'placa', 
                'data_inicio', 'data_fim', 'valor_total', 'km_rodados', 'multas_info', 'valor_restante'
            ]],
            column_config={
                "id": "ID",
                "cliente": "Cliente",
                "modelo": "Modelo",
                "placa": "Placa",
                "data_inicio": "Data Início",
                "data_fim": "Data Fim",
                "valor_total": "Valor Total",
                "km_rodados": "KM Rodados",
                "multas_info": "Multas",
                "valor_restante": "Valor Restante"
            },
            width='stretch',
            hide_index=True
        )
        
        # Seletor para ver detalhes
        st.subheader("Detalhes da Locação")
        
        # Cria uma lista de opções com ID, Cliente e Veículo
        opcoes_locacao = ["Selecione..."] + [
            f"ID: {row['id']} - Cliente: {row['cliente']} - Veículo: {row['modelo']} ({row['placa']})" 
            for _, row in df_display.iterrows()
        ]
        
        locacao_selecionada = st.selectbox(
            "Selecione uma locação para ver detalhes:",
            opcoes_locacao
        )
        
        if locacao_selecionada != "Selecione...":
            try:
                # Extrai o ID da opção selecionada
                # O formato é: "ID: 123 - Cliente: Nome - Veículo: Modelo (Placa)"
                locacao_id = int(locacao_selecionada.split(' ')[1])
                locacao = df_locacoes[df_locacoes['id'] == locacao_id].iloc[0]
            except (IndexError, ValueError) as e:
                st.error(f"Erro ao processar a seleção. Por favor, selecione uma opção válida.")
                st.stop()
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.metric("Cliente", locacao['cliente'])
                st.metric("Veículo", f"{locacao['modelo']} ({locacao['placa']})")
                st.metric("Tipo de Veículo", locacao['veiculo'])
                st.metric("Período", f"{locacao['data_inicio'].strftime('%d/%m/%Y')} a {locacao['data_fim'].strftime('%d/%m/%Y')}")
            
            with col2:
                st.metric("Valor Total", formatar_moeda(locacao['valor_total']))
                st.metric("Quilometragem", f"Saída: {locacao['km_saida']} km | Devolução: {locacao['km_volta']} km")
                st.metric("KM Rodados", f"{locacao['km_volta'] - locacao['km_saida']} km")
                st.metric("Duração", f"{max(1, (locacao['data_fim'] - locacao['data_inicio']).days)} dias")

                # Botão para gerar recibo
                if st.button("📄 Gerar Recibo", key=f"recibo_{locacao['id']}"):
                    pdf_bytes = gerar_recibo_para_download(locacao['id'])
                    if pdf_bytes:
                        st.download_button(
                            label="⬇️ Baixar Recibo",
                            data=pdf_bytes,
                            file_name=f"recibo_locacao_{locacao['id']}.pdf",
                            mime="application/pdf"
                        )
    else:
        st.info("Nenhuma locação finalizada encontrada no período selecionado.")

# 8. RELATÓRIOS (NOVA ABA)
elif menu == "Relatórios":
    st.title("📈 Relatórios")
    st.write("Aqui você pode gerar relatórios de disponibilidade da frota.")

    # Seleção de Mês e Ano
    col_mes, col_ano = st.columns(2)
    mes_selecionado = col_mes.selectbox("Selecione o Mês", range(1, 13), index=date.today().month - 1,
                                       format_func=lambda x: datetime(2000, x, 1).strftime('%B'))
    ano_selecionado = col_ano.selectbox("Selecione o Ano", range(datetime.now().year - 2, datetime.now().year + 3),
                                       index=2)

    # Obtém o primeiro e último dia do mês selecionado
    primeiro_dia_mes = date(ano_selecionado, mes_selecionado, 1)
    if mes_selecionado == 12:
        ultimo_dia_mes = date(ano_selecionado, mes_selecionado, 31)
    else:
        ultimo_dia_mes = date(ano_selecionado, mes_selecionado + 1, 1) - timedelta(days=1)

    # Otimizado: Uma única query para carros e reservas do período
    df_relatorio_data = get_relatorio_ocupacao_mensal(ano_selecionado, mes_selecionado)
    
    if df_relatorio_data.empty:
        st.warning("Nenhum veículo ativo encontrado para gerar o relatório.")
    else:
        # Separar carros únicos e reservas
        df_carros = df_relatorio_data[['id', 'modelo', 'placa']].drop_duplicates().reset_index(drop=True)
        df_reservas = df_relatorio_data[df_relatorio_data['carro_id'].notna()][['carro_id', 'data_inicio', 'data_fim', 'reserva_status']]

        # Criar a estrutura para o relatório
        # A primeira coluna será o nome do veículo, as outras serão os dias do mês
        dias_no_mes = (ultimo_dia_mes - primeiro_dia_mes).days + 1
        
        # Safety check to ensure dias_no_mes is positive
        if dias_no_mes <= 0:
            st.error(f"Erro no cálculo de dias do mês. dias_no_mes: {dias_no_mes}")
            st.error(f"primeiro_dia_mes: {primeiro_dia_mes}, ultimo_dia_mes: {ultimo_dia_mes}")
            st.stop()
            
        colunas_dias = [f"{d:02d}" for d in range(1, dias_no_mes + 1)]

        # Inicializa o DataFrame do relatório com a coluna de veículos
        df_relatorio = pd.DataFrame({'Veículo': df_carros['modelo'].fillna('') + " (" + df_carros['placa'].fillna('') + ")"})
        for dia in colunas_dias:
            df_relatorio[dia] = '' # Preenche com vazio inicialmente

        # Mapear IDs de carro para índice no df_relatorio para atualização eficiente
        carro_id_to_index = {carro_id: i for i, carro_id in enumerate(df_carros['id'])}        

        # Preencher o DataFrame do relatório com base nas reservas
        for index, row in df_reservas.iterrows():
            carro_id = row['carro_id']
            reserva_status = row['reserva_status']
            data_inicio = pd.to_datetime(row['data_inicio']).date()
            data_fim = pd.to_datetime(row['data_fim']).date()

            if carro_id in carro_id_to_index:
                idx_df_relatorio = carro_id_to_index[carro_id]
                
                for d in range(1, dias_no_mes + 1):
                    dia_atual = date(ano_selecionado, mes_selecionado, d)
                    if data_inicio <= dia_atual <= data_fim:
                        col_dia = f"{d:02d}"
                        if reserva_status == 'Reservada':
                            df_relatorio.at[idx_df_relatorio, col_dia] = 'Reservado'
                        elif reserva_status == 'Locada':
                            df_relatorio.at[idx_df_relatorio, col_dia] = 'Locado'
                        elif reserva_status == 'Finalizada':
                            df_relatorio.at[idx_df_relatorio, col_dia] = 'Finalizada'

        # Após preencher os status de reservas/locações, preencher o restante como 'Disponível'
        for r_idx in df_relatorio.index:
            for c_idx in colunas_dias:
                if df_relatorio.at[r_idx, c_idx] == '':
                    df_relatorio.at[r_idx, c_idx] = 'Disponível'

        st.dataframe(df_relatorio, hide_index=True)

        st.markdown("---")
        st.subheader("Gerar Relatório Excel")

        if st.button("📊 Gerar e Baixar Relatório de Disponibilidade", type="primary"):
            output = io.BytesIO()
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"Disponibilidade {mes_selecionado:02d}-{ano_selecionado}"

            # Obter nomes dos veículos (serão os cabeçalhos das colunas do Excel, a partir da coluna B)
            # Usa a coluna 'Veículo' do df_relatorio que já inclui Modelo e Placa
            vehicle_names_with_plate = df_relatorio['Veículo'].tolist()
            
            # Safety check for empty vehicle list
            if not vehicle_names_with_plate:
                st.error("Nenhum veículo encontrado para gerar o relatório.")
                st.stop()
            
            # Obter números dos dias (serão os cabeçalhos das linhas do Excel, a partir da linha 2)
            day_numbers_str = colunas_dias # ex: ['01', '02', ...]

            # Célula A1 vazia ou com um rótulo
            sheet.cell(row=1, column=1, value="Dia/Veículo")

            # Escrever nomes dos veículos como cabeçalhos de coluna (linha 1, começando da coluna B)
            for col_idx, vehicle_name in enumerate(vehicle_names_with_plate, start=2):
                sheet.cell(row=1, column=col_idx, value=vehicle_name)

            # Escrever números dos dias como cabeçalhos de linha (coluna 1, começando da linha 2)
            for row_idx, day_str in enumerate(day_numbers_str, start=2):
                sheet.cell(row=row_idx, column=1, value=int(day_str)) # Converte para int para exibição

            # Estilo para o cabeçalho
            header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            header_font = Font(bold=True)
            
            # Aplicar à primeira linha (cabeçalhos dos veículos)
            for col in range(1, len(vehicle_names_with_plate) + 2):
                sheet.cell(row=1, column=col).fill = header_fill
                sheet.cell(row=1, column=col).font = header_font
            
            # Aplicar à primeira coluna (cabeçalhos dos dias)
            for row in range(1, len(day_numbers_str) + 2):
                sheet.cell(row=row, column=1).fill = header_fill
                sheet.cell(row=row, column=1).font = header_font

            # Preencher dados e aplicar formatação condicional
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") # Verde
            orange_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid") # Laranja
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # Vermelho

            # Preencher as células de dados (status)
            # Iterar pelos dias (que agora são as linhas no Excel)
            for day_excel_row_idx, day_str in enumerate(day_numbers_str, start=2):
                # Obter a data atual para verificar se é um dia futuro
                dia_atual = date(ano_selecionado, mes_selecionado, int(day_str))
                hoje = date.today()
                
                # Iterar pelos veículos (que agora são as colunas no Excel)
                for vehicle_excel_col_idx, vehicle_name_full in enumerate(vehicle_names_with_plate, start=2):
                    # Encontrar o índice da linha do veículo no df_relatorio original
                    vehicle_rows = df_relatorio[df_relatorio['Veículo'] == vehicle_name_full]
                    if vehicle_rows.empty:
                        st.error(f"Veículo não encontrado no relatório: {vehicle_name_full}")
                        continue
                    
                    original_df_row_index = vehicle_rows.index[0]
                    
                    # Obter o status para o dia específico dessa linha
                    status = df_relatorio.at[original_df_row_index, day_str]
                    
                    cell = sheet.cell(row=day_excel_row_idx, column=vehicle_excel_col_idx, value=status)

                    # Aplicar cores baseadas no status
                    if status == 'Disponível':
                        cell.fill = green_fill
                    elif status == 'Reservado':
                        cell.fill = orange_fill
                    elif status == 'Locado':
                        cell.fill = red_fill
                    elif status == 'Finalizada':
                        # Para locações finalizadas, verificar se a data é futura ou passada
                        if dia_atual <= hoje:
                            cell.fill = red_fill  # Dias passados em vermelho
                        else:
                            cell.fill = green_fill  # Dias futuros em verde

            # Ajustar largura das colunas
            sheet.column_dimensions['A'].width = 10 # Largura para a coluna dos dias
            for col_idx in range(2, len(vehicle_names_with_plate) + 2):
                sheet.column_dimensions[chr(64 + col_idx)].width = 25 # Largura para os nomes dos veículos

            workbook.save(output)
            output.seek(0)

            st.download_button(
                label="Download Excel",
                data=output.getvalue(),
                file_name=f"relatorio_disponibilidade_{mes_selecionado:02d}-{ano_selecionado}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("Relatório Excel gerado com sucesso!")

# 8. GERENCIAR USUÁRIOS (APENAS ADMIN)
elif menu == "👥 Gerenciar Usuários":
    # Verificação básica de autenticação
    if not st.session_state.get('authenticated'):
        st.error("Você precisa estar autenticado para acessar esta página.")
        st.switch_page("pages/login.py")
        st.stop()

    st.title("👥 Gerenciamento de Usuários")

    tab_listar, tab_criar, tab_auditoria = st.tabs(["Listar Usuários", "Criar Usuário", "Logs de Auditoria"])

    with tab_listar:
        st.subheader("Usuários Cadastrados")

        users = supabase_auth_manager.get_users()

        if not users:
            st.info("Nenhum usuário cadastrado.")
        else:
            # Converter para DataFrame para exibição
            df_usuarios = pd.DataFrame(users)
            
            # Garantir que as colunas necessárias existam
            if 'created_at' in df_usuarios.columns:
                df_usuarios['Criado em'] = pd.to_datetime(df_usuarios['created_at']).dt.strftime('%d/%m/%Y %H:%M')
            if 'last_login' in df_usuarios.columns:
                df_usuarios['Último Login'] = pd.to_datetime(df_usuarios['last_login']).dt.strftime('%d/%m/%Y %H:%M')
            
            # Mapear valores booleanos para texto
            if 'is_active' in df_usuarios.columns:
                df_usuarios['is_active'] = df_usuarios['is_active'].map({True: 'Sim', False: 'Não', None: 'Não'})
            
            # Mapear funções para nomes amigáveis
            role_mapping = {
                'admin': 'Administrador',
                'manager': 'Gerente',
                'employee': 'Funcionário',
                'viewer': 'Visualizador'
            }
            if 'role' in df_usuarios.columns:
                df_usuarios['Função'] = df_usuarios['role'].map(role_mapping).fillna('Não definido')
            
            # Renomear colunas para exibição
            column_mapping = {
                'id': 'ID',
                'username': 'Usuário',
                'full_name': 'Nome Completo',
                'email': 'E-mail',
                'is_active': 'Ativo'
            }
            
            # Selecionar apenas as colunas que existem no DataFrame
            display_columns = []
            for col in ['ID', 'Usuário', 'Nome Completo', 'E-mail', 'Função', 'Ativo', 'Criado em', 'Último Login']:
                # Mapeia o nome de exibição para o nome da coluna original
                original_col = next((k for k, v in column_mapping.items() if v == col), col)
                if original_col in df_usuarios.columns:
                    display_columns.append(original_col)
            
            # Exibir tabela de usuários
            st.dataframe(
                df_usuarios[display_columns],
                column_config={
                    'id': 'ID',
                    'username': 'Usuário',
                    'full_name': 'Nome Completo',
                    'email': 'E-mail',
                    'is_active': 'Ativo',
                    'created_at': 'Criado em',
                    'last_login': 'Último Login',
                    'role': 'Função'
                },
                width='stretch',
                hide_index=True
            )

            # Formulário de edição
            st.subheader("Editar Usuário")
            user_options = [f"{u['id']} - {u.get('username', u.get('email', 'Sem nome'))}" for u in users]
            selected_user = st.selectbox("Selecione o usuário para editar", [""] + user_options)

            if selected_user != "Nenhum":
                user_id = int(selected_user.split(" - ")[0])
                user_data = next((u for u in users if u['id'] == user_id), None)

                if user_data:
                    st.markdown(f"**Editando:** {user_data['username']}")

                    with st.form(f"edit_user_{user_id}"):
                        col1, col2 = st.columns(2)

                        with col1:
                            new_full_name = st.text_input("Nome Completo", value=user_data['full_name'] or "")
                            new_email = st.text_input("Email", value=user_data['email'] or "")

                        with col2:
                            new_role = st.selectbox(
                                "Nível de Acesso",
                                options=list(USER_ROLES.keys()),
                                format_func=lambda x: USER_ROLES[x],
                                index=list(USER_ROLES.keys()).index(user_data['role'])
                            )
                            new_active = st.checkbox("Usuário Ativo", value=user_data['is_active'])

                        new_password = st.text_input("Nova Senha (deixe vazio para manter)", type="password")

                        if st.form_submit_button("💾 Salvar Alterações", type="primary"):
                            updates = {
                                'full_name': new_full_name,
                                'email': new_email,
                                'role': new_role,
                                'is_active': new_active
                            }

                            if new_password:
                                updates['password'] = new_password

                            success, message = supabase_auth_manager.update_user(user_id, updates)

                            if success:
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)

                    # Botão para remover usuário
                    if st.button("🗑️ Desativar Usuário", type="secondary"):
                        if user_data['username'] == current_user['username']:
                            st.error("❌ Você não pode desativar seu próprio usuário!")
                        else:
                            success, message = supabase_auth_manager.delete_user(user_id)
                            if success:
                                st.success(message)
                                st.rerun()
                            else:
                                st.error(message)
            else:
                st.info("Nenhum usuário cadastrado.")

    with tab_criar:
        st.subheader("Criar Novo Usuário")

        with st.form("create_user_form"):
            col1, col2 = st.columns(2)

            with col1:
                username = st.text_input("Nome de Usuário", help="Nome único para login")
                full_name = st.text_input("Nome Completo")
                email = st.text_input("Email")

            with col2:
                password = st.text_input("Senha", type="password", help="Mínimo 6 caracteres")
                confirm_password = st.text_input("Confirmar Senha", type="password")
                role = st.selectbox(
                    "Nível de Acesso",
                    options=list(USER_ROLES.keys()),
                    format_func=lambda x: USER_ROLES[x]
                )

            if st.form_submit_button("👤 Criar Usuário", type="primary"):
                if not username or not password or not full_name:
                    st.error("❌ Preencha todos os campos obrigatórios!")
                elif password != confirm_password:
                    st.error("❌ As senhas não coincidem!")
                elif len(password) < 6:
                    st.error("❌ A senha deve ter pelo menos 6 caracteres!")
                else:
                    success, message = supabase_auth_manager.create_user(
                        username, password, role, full_name, email
                    )

                    if success:
                        st.success(message)
                        st.balloons()
                        st.rerun()
                    else:
                        st.error(message)

    with tab_auditoria:
        st.subheader("Logs de Auditoria")

        logs = supabase_auth_manager.get_audit_logs(200)  # Últimos 200 registros

        if logs:
            df_logs = pd.DataFrame(logs)
            df_logs['timestamp'] = pd.to_datetime(df_logs['timestamp']).dt.strftime('%d/%m/%Y %H:%M:%S')

            st.dataframe(
                df_logs,
                column_config={
                    'timestamp': 'Data/Hora',
                    'username': 'Usuário',
                    'action': 'Ação',
                    'resource': 'Recurso',
                    'details': 'Detalhes',
                    'ip_address': 'IP'
                },
                width='stretch'
            )

            # Resumo de atividades
            st.markdown("---")
            st.subheader("📊 Resumo de Atividades")

            col1, col2, col3 = st.columns(3)

            with col1:
                total_logins = len([l for l in logs if l['action'] == 'login'])
                st.metric("Total de Logins", total_logins)

            with col2:
                total_users_created = len([l for l in logs if l['action'] == 'user_created'])
                st.metric("Usuários Criados", total_users_created)

            with col3:
                active_users = len(set([l['username'] for l in logs if l['username']]))
                st.metric("Usuários Ativos", active_users)
        else:
            st.info("Nenhum log de auditoria encontrado.")

# 9. GERENCIAMENTO DE MULTAS (NOVA ABA)
elif menu == "Gerenciar Multas":
    render_section_header(
        title="Gerenciamento de Multas",
        subtitle="Identifique rapidamente o veículo multado, registre o ocorrido e acompanhe o status.",
        icon="🚨",
        trail=["Operação", "Multas"]
    )

    if 'multas_data_consulta' not in st.session_state:
        st.session_state.multas_data_consulta = datetime.now().date()
    if 'multas_carros_result' not in st.session_state:
        st.session_state.multas_carros_result = None

    st.subheader("1️⃣ Localizar veículo na data da infração")
    with st.form("form_consulta_multas"):
        col_f1, col_f2 = st.columns([2, 1])
        with col_f1:
            data_multa = st.date_input(
                "Data da infração",
                value=st.session_state.multas_data_consulta,
                key="multa_data_input"
            )
        with col_f2:
            st.markdown("&nbsp;")
            consultar = st.form_submit_button("🔍 Consultar veículos", width='stretch')

    if consultar:
        resultado = run_query(
            """
            SELECT DISTINCT c.id, c.marca, c.modelo, c.placa, 
                            cl.nome, cl.cpf, cl.cnh, r.id AS reserva_id,
                            r.data_inicio, r.data_fim
            FROM carros c
            JOIN reservas r ON c.id = r.carro_id
            JOIN clientes cl ON r.cliente_id = cl.id
            WHERE (%s::date BETWEEN DATE(r.data_inicio) AND DATE(r.data_fim))
              AND r.reserva_status IN ('Locada', 'Reservada', 'Finalizada')
            ORDER BY r.data_inicio
            """,
            (data_multa,),
            fetch=True
        )
        if isinstance(resultado, str):
            st.error(f"Erro ao buscar veículos: {resultado}")
        else:
            st.session_state.multas_data_consulta = data_multa
            st.session_state.multas_carros_result = resultado

    carros_df = st.session_state.get('multas_carros_result')

    if carros_df is not None and not isinstance(carros_df, str) and not carros_df.empty:
        st.success(f"{len(carros_df)} veículo(s) estavam locados em {st.session_state.multas_data_consulta:%d/%m/%Y}.")
        opcoes = [None] + carros_df.index.tolist()
        selecao = st.selectbox(
            "Selecione o veículo para registrar a multa",
            opcoes,
            index=0,
            format_func=lambda idx: "Selecione o veículo..." if idx is None else (
                f"ID {carros_df.at[idx, 'reserva_id']} • "
                f"{carros_df.at[idx, 'marca']} {carros_df.at[idx, 'modelo']} "
                f"({carros_df.at[idx, 'placa']}) – "
                f"{pd.to_datetime(carros_df.at[idx, 'data_inicio']).date():%d/%m} → "
                f"{pd.to_datetime(carros_df.at[idx, 'data_fim']).date():%d/%m}"
            ),
            key="multas_select_veiculo"
        )

        if selecao is not None:
            carro = carros_df.loc[selecao]
            resumo_cols = st.columns(3)
            resumo_cols[0].metric("Cliente", carro['nome'])
            resumo_cols[1].metric("Placa", carro['placa'])
            resumo_cols[2].metric(
                "Período da locação",
                f"{pd.to_datetime(carro['data_inicio']).date():%d/%m} → {pd.to_datetime(carro['data_fim']).date():%d/%m}"
            )

            st.caption(f"CPF: {carro['cpf']} • CNH: {carro['cnh']} • Reserva #{int(carro['reserva_id'])}")
            st.markdown("---")
            st.subheader("2️⃣ Registrar multa")

            with st.form("form_registro_multa"):
                col_reg1, col_reg2 = st.columns(2)
                with col_reg1:
                    tipo_multa = st.selectbox(
                        "Tipo de infração",
                        [
                            "Excesso de Velocidade",
                            "Estacionamento em Local Proibido",
                            "Avanço de Sinal",
                            "Uso de Celular ao Volante",
                            "Outra Infração de Trânsito"
                        ],
                        key="campo_tipo_multa"
                    )
                    valor_multa = st.number_input(
                        "Valor da multa (R$)",
                        min_value=0.0,
                        step=10.0,
                        value=200.0,
                        key="campo_valor_multa"
                    )
                with col_reg2:
                    data_hora_infracao = st.datetime_input(
                        "Data e hora da infração",
                        value=datetime.combine(st.session_state.multas_data_consulta, datetime.now().time()),
                        key="campo_datahora_multa"
                    )
                    local_infracao = st.text_input(
                        "Local da infração (opcional)",
                        key="campo_local_multa"
                    )

                observacao = st.text_area("Observações adicionais (opcional)", key="campo_observacao_multa")
                submit_multa = st.form_submit_button("🚨 Registrar multa", width='stretch')

                if submit_multa:
                    try:
                        run_query(
                            """
                            INSERT INTO multas (reserva_id, tipo, valor, data_multa, status, local_infracao, observacao)
                            VALUES (%s, %s, %s, %s, 'Pendente', %s, %s)
                            """,
                            (carro['reserva_id'], tipo_multa, valor_multa, data_hora_infracao, local_infracao or None, observacao or None)
                        )
                        run_query(
                            "UPDATE reservas SET status = 'Com Multa Pendente' WHERE id = %s",
                            (carro['reserva_id'],)
                        )
                        st.toast("Multa registrada com sucesso!", icon="✅")
                        st.success("✅ Multa registrada e reserva atualizada.")
                        st.session_state.multas_carros_result = None
                        #st.session_state.multas_select_veiculo = None
                        time.sleep(1)
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Erro ao registrar multa: {e}")
    else:
        st.info("Consulte uma data para listar os veículos locados e registrar a multa correspondente.")

    st.divider()
    st.subheader("3️⃣ Multas registradas")

    hoje = datetime.now().date()
    col_filtro_inicio, col_filtro_fim = st.columns(2)
    data_inicio = col_filtro_inicio.date_input("Data inicial", value=hoje - timedelta(days=30), key="multa_filtro_inicio")
    data_fim = col_filtro_fim.date_input("Data final", value=hoje, key="multa_filtro_fim")

    multas_df = run_query(
        """
        SELECT m.*, 
               c.placa, c.modelo,
               cl.nome AS cliente_nome, cl.cpf AS cliente_cpf, cl.cnh AS cliente_cnh,
               cl.telefone AS cliente_telefone, cl.endereco AS cliente_endereco
        FROM multas m
        JOIN reservas r ON m.reserva_id = r.id
        JOIN carros c ON r.carro_id = c.id
        JOIN clientes cl ON r.cliente_id = cl.id
        WHERE DATE(m.data_multa) BETWEEN %s AND %s
        ORDER BY m.data_multa DESC
        """,
        (data_inicio, data_fim + timedelta(days=1)),
        fetch=True
    )

    if isinstance(multas_df, str):
        st.error(f"Erro ao buscar multas: {multas_df}")
    elif multas_df.empty:
        st.info("Nenhuma multa encontrada para o período selecionado.")
    else:
        total_multas = len(multas_df)
        valor_pendente = multas_df.loc[multas_df['status'] == 'Pendente', 'valor'].sum()
        valor_pago = multas_df.loc[multas_df['status'] == 'Paga', 'valor'].sum()

        metric_cols = st.columns(3)
        metric_cols[0].metric("Total de multas", total_multas)
        metric_cols[1].metric("Valor pendente", formatar_moeda(valor_pendente))
        metric_cols[2].metric("Valor pago", formatar_moeda(valor_pago))

        status_opcoes = ['Pendente', 'Paga', 'Isentada']
        status_badge = {
            'Pendente': '🔴 Pendente',
            'Paga': '🟢 Paga',
            'Isentada': '🔵 Isentada'
        }

        for _, multa in multas_df.iterrows():
            data_formatada = pd.to_datetime(multa['data_multa']).strftime("%d/%m/%Y %H:%M")
            with st.container(border=True):
                st.markdown(f"### 🎫 Multa #{multa['id']} — {status_badge.get(multa['status'], multa['status'])}")
                info_cols = st.columns([1.3, 1.3, 1])
                with info_cols[0]:
                    st.markdown("**Locatário**")
                    st.write(f"Nome: {multa['cliente_nome']}")
                    st.write(f"CPF: {multa.get('cliente_cpf', '—')}")
                    st.write(f"CNH: {multa.get('cliente_cnh', '—')}")
                    st.write(f"Telefone: {multa.get('cliente_telefone', '—')}")
                    st.write(f"Endereço: {multa.get('cliente_endereco', '—')}")
                with info_cols[1]:
                    st.markdown("**Veículo/Reserva**")
                    st.write(f"Modelo: {multa['modelo']}")
                    st.write(f"Placa: {multa['placa']}")
                    st.write(f"Reserva: #{multa['reserva_id']}")
                    st.write(f"Data da infração: {data_formatada}")
                    st.write(f"Tipo: {multa['tipo']}")
                    if pd.notna(multa['local_infracao']):
                        st.write(f"Local: {multa['local_infracao']}")
                with info_cols[2]:
                    st.metric("Valor", formatar_moeda(multa['valor']))
                    with st.form(f"form_status_multa_{multa['id']}"):
                        novo_status = st.selectbox(
                            "Status",
                            status_opcoes,
                            index=status_opcoes.index(multa['status']) if multa['status'] in status_opcoes else 0,
                            key=f"status_select_{multa['id']}"
                        )
                        salvar_status = st.form_submit_button("Atualizar status")
                        if salvar_status:
                            if novo_status != multa['status']:
                                try:
                                    run_query("UPDATE multas SET status=%s WHERE id=%s", (novo_status, multa['id']))
                                    
                                    # Atualizar status da reserva baseado no novo status da multa
                                    if novo_status == "Paga":
                                        run_query("UPDATE reservas SET status='Finalizada' WHERE id=%s", (multa['reserva_id'],))
                                    else:
                                        run_query("UPDATE reservas SET status='Com Multa Pendente' WHERE id=%s", (multa['reserva_id'],))
    
                                    st.toast("Status atualizado!", icon="✅")
                                    st.success("Status da multa atualizado com sucesso.")
                                    time.sleep(1)
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"Erro ao atualizar status: {e}")
                            else:
                                st.info("O status selecionado já está aplicado.")

                if pd.notna(multa['observacao']) and multa['observacao'].strip():
                    with st.expander("Observações"):
                        st.write(multa['observacao'])

# Chamar a função main() quando o Dashboard estiver selecionado
if st.session_state.get('main_menu_selector') == "Dashboard":
    try:
        main()
    except Exception as e:
        st.error(f"Ocorreu um erro ao carregar o dashboard: {str(e)}")
